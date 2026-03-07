# coding=utf-8
"""
EpiRank GUI — Epidemic Risk Analysis System (PySide6)

Implements the EpiRank algorithm described in:
    Chung-Yuan Huang et al., "EpiRank: Modeling Bidirectional Disease Spread
    in Asymmetric Commuting Networks", 2014.

The EpiRank model estimates the relative epidemic risk of 353 townships in
Taiwan by combining a bidirectional commuting network with a PageRank-like
iterative algorithm.  The core formula (Eq. 4–6 of the paper) is:

    ER = (1 − d) · exFac  +  d · (daytime · W^T · ER  +  (1 − daytime) · W · ER)

where
    ER       = epidemic risk vector (N×1), sums to 1.0 at convergence
    d        = damping factor (default 0.95); higher → network dominates
    daytime  = forward/backward movement weight (0–1);
               0.0 = backward only (evening commute home → residence),
               0.5 = bidirectional (equal weight),
               1.0 = forward only (morning commute home → workplace)
    W        = column-normalised OD (origin-destination) commuting matrix
    W^T      = column-normalised transpose of the raw OD matrix
    exFac    = external factor vector (default uniform 1/N)

Classification uses the head/tail breaks method (Jiang 2013), recursively
splitting at the mean three times to produce four levels:
    NC (non-core) → C-III → C-II → C-I (highest risk).

The GUI reproduces all key figures/tables from the paper:
    Tab  0  Results Table           – ranked EpiRank scores for all townships
    Tab  1  Network Map             – commuting network visualisation
    Tab  2  Core Classification     – Table 1 (head/tail break counts by method)
    Tab  3  Correlations            – Table 2 (Pearson/Spearman/Recall/Precision)
    Tab  4  Commuter Flow           – Figure 2  (7 sub-plots: map, scatter, hist)
    Tab  5  Frequency Distributions – Figure 3  (disease frequency + log ratio)
    Tab  6  Frequency Distribution  – Figure 6  (EpiRank freq. by daytime)
    Tab  7  EpiRank vs Disease      – Figure 9  (stacked % bars)
    Tab  8  Index Comparison        – Figure 10 (EpiRank vs PageRank vs HITS)
    Tab  9  Disease Map             – Figure 4  (spatial disease severity)
    Tab 10  EpiRank Map             – Figure 7  (spatial EpiRank levels)
    Tab 11  EpiRank vs Disease Map  – Figure 8  (overlay: prediction vs actual)
    Tab 12  Log                     – computation log
    Tab 13  Sensitivity Analysis    – Figure 11 (daytime × d heatmaps)

Modernised from ERA.py (Python 2.7, 崇源) → Python 3.13 / PySide6 / NumPy.
"""

import sys
import os
import numpy as np
import networkx as nx
from scipy import stats as st
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment

from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QGroupBox, QLabel, QDoubleSpinBox, QSpinBox, QPushButton,
    QFileDialog, QTextEdit, QTabWidget, QProgressBar,
    QTableWidget, QTableWidgetItem, QMessageBox,
    QFormLayout, QComboBox, QStatusBar
)
from PySide6.QtCore import Qt, QThread, Signal
from PySide6.QtGui import QFont, QColor, QAction

import matplotlib
from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.backends.backend_qtagg import NavigationToolbar2QT as NavigationToolbar
from matplotlib.figure import Figure
from matplotlib import font_manager

# ---- Configure Matplotlib for CJK (Traditional Chinese) support ----
# Try fonts in order of preference on macOS
_CJK_FONT_CANDIDATES = [
    'Heiti TC',         # macOS built-in Traditional Chinese
    'PingFang HK',      # macOS PingFang
    'Arial Unicode MS', # Wide Unicode coverage
    'Hiragino Sans',    # macOS Hiragino
    'Noto Sans CJK TC', # Google Noto (if installed)
]

_cjk_font_found = None
_available_font_names = {f.name for f in font_manager.fontManager.ttflist}
for _candidate in _CJK_FONT_CANDIDATES:
    if _candidate in _available_font_names:
        _cjk_font_found = _candidate
        break

if _cjk_font_found:
    matplotlib.rcParams['font.sans-serif'] = [_cjk_font_found] + matplotlib.rcParams.get('font.sans-serif', [])
    matplotlib.rcParams['font.family'] = 'sans-serif'
    matplotlib.rcParams['axes.unicode_minus'] = False   # Fix minus sign display
    print(f"[Matplotlib] Using CJK font: {_cjk_font_found}")
else:
    print("[Matplotlib] WARNING: No CJK font found. Chinese characters may not display correctly.")

# ============================================================
# Constants — dictionary keys for town_data records
# Each township (鄉鎮市) is stored as town_data[db_ID] = {key: value, ...}
# ============================================================
KEY_POST_CODE            = 'post_code'
KEY_DB_ID                = 'db_ID'
KEY_COUNTY               = 'county'
KEY_TOWN                 = 'town'
KEY_POS_XY               = 'pos_xy'
KEY_POPULATION           = 'population'
KEY_AREA                 = 'area'
KEY_DENSITY              = 'density'
KEY_NORMALIZED_DENSITY   = 'normalized_density'
KEY_AGE_0_14             = 'age_0_14'
KEY_AGE_15_64            = 'age_15_64'
KEY_AGE_65               = 'age_65'
KEY_LOCAL_COMMUTER_TYPE1 = 'local_commuter_type1'
KEY_OUT_COMMUTER_TYPE1   = 'out_commuter_type1'
KEY_IN_COMMUTER_TYPE1    = 'in_commuter_type1'
KEY_COMMUTER_TYPE1       = 'commuter_type1'
KEY_RAILROAD_ZONE        = 'railroad_zone'
KEY_FLU_TOTAL_CASES      = 'flu_total_cases'
KEY_EV_AVERAGE_CASES     = 'EV_average_cases'
KEY_SARS_TOTAL_CASES     = 'sars_total_cases'

# Greater Taipei Metropolitan Area (大台北都會區): 48 townships by db_ID.
# Used for regional zoom views in Figures 7–8 and gTaipei SARS correlation
# analysis.  The set covers Taipei City (台北市), New Taipei City (新北市),
# and Keelung City (基隆市) administrative divisions.
GTAIPEI_DB_IDS = set(range(0, 29)) | set(range(303, 310)) | set(range(330, 397))

# ============================================================
# Core EpiRank Engine (modernized from ERA.py)
# ============================================================
#
# 演算法總覽 (Algorithm Overview)
# ──────────────────────────────
# EpiRank 的核心洞見：傳染病的擴散並非隨機，而是沿著人類每日通勤的
# 路徑流動。每天早晨，數百萬人從住家（origin）移動到工作地（destination），
# 晚間再返回——這條雙向的人流，就是疾病傳播的高速公路。
#
# The key insight of EpiRank: epidemic spread is not random — it flows
# along the daily commuting paths of millions of people.  Every morning,
# commuters move from home (origin) to work (destination); every evening,
# they return.  This bidirectional human flow is the highway of disease
# transmission.
#
# EpiRank 借鏡 Google PageRank 的精神：一個網頁的重要性取決於「誰連結
# 到它」；同理，一個鄉鎮的疫情風險取決於「誰通勤到這裡、誰從這裡回家」。
# 但 EpiRank 比 PageRank 更進一步——它同時考慮了「去程」（早晨通勤，
# 人口從住家擴散到工作地）和「回程」（晚間通勤，人口從工作地回流到住家），
# 用 daytime 參數控制兩個方向的權重。
#
# EpiRank borrows from Google's PageRank philosophy: a webpage's
# importance depends on "who links to it"; analogously, a township's
# epidemic risk depends on "who commutes here and who returns home from
# here."  But EpiRank goes further — it simultaneously models the
# *forward trip* (morning: population spreads from home to work) and the
# *return trip* (evening: population flows back from work to home), with
# the ``daytime`` parameter controlling the balance between the two.
#
# 三段式建構流程 (Three-Stage Construction)
# ──────────────────────────────────────────
# Stage 1 — 建網 (Network Construction)
#   從人口普查通勤 OD 矩陣建立 353 節點的有向圖。
#   → build_commuting_network()
#
# Stage 2 — 正規化 (Matrix Normalisation)
#   將原始通勤矩陣分別做 column-normalisation，得到兩個隨機矩陣：
#     W   = 原始 OD 矩陣 column-normalised   → 模擬「回家」方向
#     W^T = OD 轉置矩陣 column-normalised    → 模擬「去上班」方向
#   → compute_epidemic_risk() 前半段
#
# Stage 3 — 迭代收斂 (Iterative Convergence)
#   反覆套用 EpiRank 公式直到風險向量穩定：
#     ER = (1-d)·(1/N) + d·[daytime·W^T·ER + (1-daytime)·W·ER]
#   收斂後 ER 向量加總為 1.0，每個元素代表該鄉鎮的相對疫情風險。
#   → compute_epidemic_risk() 後半段
#
# 為什麼這很優美？(Why is this elegant?)
# ────────────────────────────────────────
# 1. 只需一個通勤 OD 矩陣，就能預測三種不同疾病（流感、腸病毒、SARS）
#    的空間分布——說明通勤結構本身就是疫情風險的根本驅動力。
# 2. daytime 參數讓模型能區分「白天型」和「夜間型」的傳播路徑：
#    daytime=0.0 → 純回程（疫情跟著人回家散播到住宅區）
#    daytime=0.5 → 雙向等權（最符合現實）
#    daytime=1.0 → 純去程（疫情在工作地累積）
# 3. 數學上保證收斂（column-stochastic matrix 的 power iteration），
#    且收斂速度由 damping factor d 控制。
# ============================================================

def sorted_map(mapping):
    """Sort a dictionary by value descending, then by key ascending.

    Used to rank townships by EpiRank score for the Results Table (Tab 0)
    and Excel auto-save output.
    """
    return sorted(mapping.items(), key=lambda kv: (-kv[1], kv[0]))


def build_basic_table_of_towns(town_data, path='bs.xlsx', sheet='town_data',
                                number_of_sub_towns=409, row_base=2):
    """Load basic township metadata (population, area, density, age structure)
    from bs.xlsx.

    The spreadsheet contains 409 sub-township rows which are aggregated into
    353 unique townships.  Only the first row per town_name is kept, which
    provides the merged township-level statistics.
    """
    wb = load_workbook(path, data_only=True)
    s = wb[sheet]
    old_town = None

    for row_idx in range(number_of_sub_towns):
        r = row_idx + row_base

        # ── 防禦：跳過含空白儲存格的列，避免 int(None)/float(None) 崩潰 ──
        cell_val = s.cell(row=r, column=1).value
        if cell_val is None:
            continue
        db_ID              = int(cell_val)
        county             = s.cell(row=r, column=2).value
        town               = s.cell(row=r, column=3).value
        pos_xy             = (round(float(s.cell(row=r, column=7).value or 0), 2),
                              round(float(s.cell(row=r, column=8).value or 0), 2))
        raw_population     = s.cell(row=r, column=9).value or 0
        sub_percentage     = float(s.cell(row=r, column=10).value or 0)
        area               = float(s.cell(row=r, column=12).value or 0)
        density            = float(s.cell(row=r, column=13).value or 0)
        normalized_density = float(s.cell(row=r, column=14).value or 0)
        age_0_14           = float(s.cell(row=r, column=15).value or 0)
        age_15_64          = float(s.cell(row=r, column=16).value or 0)
        age_65             = float(s.cell(row=r, column=17).value or 0)
        
        # 防禦：sub_percentage 為 0 時以 raw_population 代替，避免除以零
        population         = (float(raw_population / sub_percentage)
                              if sub_percentage > 0 else float(raw_population))
        if town != old_town:
            town_data[db_ID] = {
                KEY_COUNTY: county, KEY_TOWN: town, KEY_POS_XY: pos_xy,
                KEY_POPULATION: population, KEY_AREA: area,
                KEY_DENSITY: density, KEY_NORMALIZED_DENSITY: normalized_density,
                KEY_AGE_0_14: age_0_14, KEY_AGE_15_64: age_15_64, KEY_AGE_65: age_65,
                KEY_LOCAL_COMMUTER_TYPE1: 0, KEY_OUT_COMMUTER_TYPE1: 0,
                KEY_IN_COMMUTER_TYPE1: 0, KEY_RAILROAD_ZONE: 0
            }
        old_town = town
    wb.close()


def build_flu_reported_cases(town_data, path='Flu.xlsx', sheet='2009',
                              number_of_towns=353, row_base=2):
    """Load 2009 influenza case counts from Flu.xlsx into town_data.

    Each township is matched by (county, town) name pair.
    Source: Taiwan CDC (疾管署) yearly surveillance data.
    """
    wb = load_workbook(path, data_only=True)
    s = wb[sheet]
    check_list = {}
    for row_idx in range(number_of_towns):
        r = row_idx + row_base
        county = s.cell(row=r, column=1).value
        town_name = s.cell(row=r, column=2).value
        if not check_list:
            for db_ID in town_data.keys():
                check_list[(town_data[db_ID][KEY_COUNTY], town_data[db_ID][KEY_TOWN])] = db_ID
        db_ID = check_list.get((county, town_name), None)
        if db_ID is not None:
            raw = s.cell(row=r, column=3).value
            town_data[db_ID][KEY_FLU_TOTAL_CASES] = int(raw) if raw is not None else 0
    wb.close()


def build_ev_reported_cases(town_data, path='ev.xlsx', sheet='2000_2008',
                             number_of_towns=353, row_base=2):
    """Load 2000–2008 average enterovirus case counts from ev.xlsx into
    town_data.  EV data is stored as a float (yearly average)."""
    wb = load_workbook(path, data_only=True)
    s = wb[sheet]
    check_list = {}
    for row_idx in range(number_of_towns):
        r = row_idx + row_base
        county = s.cell(row=r, column=1).value
        town_name = s.cell(row=r, column=2).value
        if not check_list:
            for db_ID in town_data.keys():
                check_list[(town_data[db_ID][KEY_COUNTY], town_data[db_ID][KEY_TOWN])] = db_ID
        db_ID = check_list.get((county, town_name), None)
        if db_ID is not None:
            raw = s.cell(row=r, column=3).value
            town_data[db_ID][KEY_EV_AVERAGE_CASES] = float(raw) if raw is not None else 0.0
    wb.close()


def build_sars_reported_cases(town_data, path='SARS.xlsx', sheet='2003',
                               number_of_towns=353, row_base=2):
    """Load 2003 SARS case counts from SARS.xlsx into town_data.

    SARS data is used for Greater Taipei correlation analysis only (not
    for the full-Taiwan EpiRank vs disease comparison in the paper).
    """
    wb = load_workbook(path, data_only=True)
    s = wb[sheet]
    check_list = {}
    for row_idx in range(number_of_towns):
        r = row_idx + row_base
        county = s.cell(row=r, column=1).value
        town_name = s.cell(row=r, column=2).value
        if not check_list:
            for db_ID in town_data.keys():
                check_list[(town_data[db_ID][KEY_COUNTY], town_data[db_ID][KEY_TOWN])] = db_ID
        db_ID = check_list.get((county, town_name), None)
        if db_ID is not None:
            raw = s.cell(row=r, column=3).value
            town_data[db_ID][KEY_SARS_TOTAL_CASES] = int(raw) if raw is not None else 0
    wb.close()


def build_commuting_network(g, town_data, path='cn.xlsx', sheet='353C',
                             number_of_towns=353, row_base=6, col_base=6):
    """Build the 353×353 directed commuting network from cn.xlsx.

    This is Stage 1 of the EpiRank pipeline — constructing the commuting
    graph that serves as the "skeleton" for disease transmission modelling.

    The OD (Origin-Destination) matrix in the spreadsheet encodes the
    daily commuting patterns from the 2000 Taiwan population census:

        OD[i][j] = number of commuters living in township i
                   who work in township j

    Key properties of this network:
    - Directed: commuting from A→B does not imply B→A
    - Weighted: edge weight = commuter count (not binary)
    - Self-loops: OD[i][i] = local commuters who live and work in the
      same township (typically ~84% of all commuters, per the paper)
    - Asymmetric: a bedroom suburb may send 50,000 commuters to the city
      centre but receive only 2,000 in return

    The resulting DiGraph has ~353 nodes and ~15,000+ directed edges.
    Each node carries geographic coordinates (TWD97 TM2, in metres)
    for spatial plotting.

    Paper Section II.A: "We obtain the number of commuting trips between
    each pair of townships from the 2000 population census."

    Note: ``read_only=False`` is used intentionally.  The ``read_only=True``
    mode uses a streaming XML parser that makes random-access .cell(row, col)
    calls O(n²)-slow, causing this function to take >100 s instead of <1 s.
    """
    wb = load_workbook(path, read_only=False, data_only=True)
    s = wb[sheet]
    cache = {}

    for row_idx in range(number_of_towns):
        r = row_idx + row_base
        raw_seq = s.cell(row=r, column=1).value
        if raw_seq is None:
            continue  # 防禦：跳過空白列
        row_seq_no = int(raw_seq)
        if row_seq_no in cache:
            row_code, row_db_ID = cache[row_seq_no]
        else:
            row_code = str(s.cell(row=r, column=2).value or '')
            raw_db = s.cell(row=r, column=3).value
            if raw_db is None:
                continue
            row_db_ID = int(raw_db)
            cache[row_seq_no] = (row_code, row_db_ID)

        for col_idx in range(number_of_towns):
            c = col_idx + col_base
            raw_col_seq = s.cell(row=1, column=c).value
            if raw_col_seq is None:
                continue
            col_seq_no = int(raw_col_seq)
            if col_seq_no in cache:
                col_code, col_db_ID = cache[col_seq_no]
            else:
                col_code = str(s.cell(row=2, column=c).value or '')
                raw_col_db = s.cell(row=3, column=c).value
                if raw_col_db is None:
                    continue
                col_db_ID = int(raw_col_db)
                cache[col_seq_no] = (col_code, col_db_ID)

            raw_commuters = s.cell(row=r, column=c).value
            commuters = int(raw_commuters) if raw_commuters is not None else 0

            if commuters > 0:
                # 防禦：確保 db_ID 存在於 town_data 中
                if row_db_ID not in town_data or col_db_ID not in town_data:
                    continue
                if not g.has_node(row_seq_no):
                    g.add_node(row_seq_no, post_code=row_code, db_ID=row_db_ID,
                               posx=town_data[row_db_ID][KEY_POS_XY][0],
                               posy=town_data[row_db_ID][KEY_POS_XY][1])
                if not g.has_node(col_seq_no):
                    g.add_node(col_seq_no, post_code=col_code, db_ID=col_db_ID,
                               posx=town_data[col_db_ID][KEY_POS_XY][0],
                               posy=town_data[col_db_ID][KEY_POS_XY][1])

                g.add_edge(row_seq_no, col_seq_no, weight=float(commuters),
                           commuter_type1=float(commuters))

                town_data[row_db_ID][KEY_OUT_COMMUTER_TYPE1] += commuters
                town_data[col_db_ID][KEY_IN_COMMUTER_TYPE1] += commuters
    wb.close()


def get_pearson_cor(dic1, dic2):
    """Pearson correlation between two dicts sharing the same keys.

    Used in Table 2 and Figure 11 sensitivity analysis to measure the
    linear association between network indices and disease case counts.

    前提：dic1 與 dic2 的 key 集合必須完全相同（皆來自同一個 nodes 清單）。
    Precondition: dic1 and dic2 must share identical key sets (both are
    built from the same ``nodes`` list in ``ComputeWorker.run``).
    """
    assert set(dic1.keys()) == set(dic2.keys()), \
        f"Key mismatch: {len(dic1)} vs {len(dic2)} keys"
    keys = list(dic1.keys())
    if len(keys) < 3:
        return (float('nan'), float('nan'))
    n1 = [dic1[k] for k in keys]
    n2 = [dic2[k] for k in keys]
    r, p = st.pearsonr(n1, n2)
    return round(r, 6), round(p, 6)


def get_spearman_cor(dic1, dic2):
    """Spearman rank correlation between two dicts sharing the same keys.

    Spearman is rank-based, so it captures monotonic (not necessarily
    linear) relationships — more robust to outliers than Pearson.

    前提：dic1 與 dic2 的 key 集合必須完全相同。
    Precondition: dic1 and dic2 must share identical key sets.
    """
    assert set(dic1.keys()) == set(dic2.keys()), \
        f"Key mismatch: {len(dic1)} vs {len(dic2)} keys"
    keys = list(dic1.keys())
    if len(keys) < 3:
        return (float('nan'), float('nan'))
    n1 = [dic1[k] for k in keys]
    n2 = [dic2[k] for k in keys]
    r, p = st.spearmanr(n1, n2)
    return round(r, 6), round(p, 6)


def get_kendalltau_cor(dic1, dic2):
    """Kendall's tau rank correlation between two dicts sharing the same keys.

    Used in the Excel auto-save for SARS vs network index correlations.

    前提：dic1 與 dic2 的 key 集合必須完全相同。
    Precondition: dic1 and dic2 must share identical key sets.
    """
    assert set(dic1.keys()) == set(dic2.keys()), \
        f"Key mismatch: {len(dic1)} vs {len(dic2)} keys"
    keys = list(dic1.keys())
    if len(keys) < 3:
        return (float('nan'), float('nan'))
    n1 = [dic1[k] for k in keys]
    n2 = [dic2[k] for k in keys]
    r, p = st.kendalltau(n1, n2)
    return round(r, 6), round(p, 6)


# ============================================================
# Head/Tail Breaks Classification
# ============================================================
# EpiRank 計算完成後，我們需要一個方法將連續的風險分數轉化為離散的
# 疫情嚴重度等級。本文選用 head/tail breaks（Jiang 2013）——一個
# 專為重尾分佈（heavy-tailed distribution）設計的自然分類方法。
#
# After EpiRank produces a continuous risk score for each township, we
# need a method to convert these scores into discrete severity levels.
# The paper uses head/tail breaks (Jiang 2013) — a classification
# method specifically designed for heavy-tailed distributions.
#
# 為什麼不用等距分類或分位數？因為 EpiRank 分數（以及疾病案例數）
# 呈現典型的重尾分布：大多數鄉鎮風險很低，少數鄉鎮風險極高。
# 等距分類會把幾乎所有鄉鎮歸為同一級；分位數則會強制各級人數相等，
# 忽略資料本身的自然斷點。Head/tail breaks 讓資料「自己說話」：
#
# Why not equal-interval or quantile classification?  Because EpiRank
# scores (and disease case counts) follow a heavy-tailed distribution:
# most townships have very low risk, a few have extremely high risk.
# Equal-interval would lump nearly all townships into one class;
# quantiles would force equal counts per class, ignoring natural breaks
# in the data.  Head/tail breaks lets the data "speak for itself":
#
# 遞迴分裂過程 (Recursive splitting):
#
#   Round 1: 全部 353 townships
#     ├─ tail (≤ mean₁): ~239 townships → NC (non-core)
#     └─ head (> mean₁): ~114 townships    ← 再分
#         Round 2:
#         ├─ tail (≤ mean₂): ~67 townships → C-III
#         └─ head (> mean₂): ~47 townships    ← 再分
#             Round 3:
#             ├─ tail (≤ mean₃): ~31 townships → C-II
#             └─ head (> mean₃): ~16 townships → C-I (highest risk)
#
# 三次分裂 → 四個等級：NC → C-III → C-II → C-I
# 每一次分裂都在「少數高值」與「多數低值」之間找到自然斷點。
#
# Paper Section III.A: "we need to identify the core and non-core
# townships based on a given index.  Head/tail breaks recursively
# partition a dataset by its mean."
# ============================================================
LEVEL_COLORS = {
    'NC':    '#3a8f3e',   # muted green  — non-core
    'C-III': '#c8b840',   # olive-yellow — core level III
    'C-II':  '#e07830',   # orange       — core level II
    'C-I':   '#cc2020',   # dark red     — core level I (highest)
}
LEVEL_ORDER = ['NC', 'C-III', 'C-II', 'C-I']  # drawing order: background → foreground


def head_tail_breaks(values, n_breaks=3):
    """Recursively split at the mean (head/tail breaks, Jiang 2013).

    Parameters
    ----------
    values : array-like
        Numeric values to classify (e.g. EpiRank scores, disease counts).
    n_breaks : int
        Number of recursive splits.  Default 3 → 4 output levels.

    Returns
    -------
    list[float]
        Sorted list of break points (ascending).  With n_breaks=3 the
        result has exactly 3 values: [mean_all, mean_head1, mean_head2].
    """
    breaks = []
    head = values.copy()
    for _ in range(n_breaks):
        if len(head) < 2:
            break
        m = head.mean()
        breaks.append(m)
        head = head[head > m]
    return sorted(breaks)


def classify_by_breaks(values, breaks):
    """Classify values into NC / C-III / C-II / C-I based on head/tail breaks.

    The boundary rule uses ``<=`` for the lower levels, so a value exactly
    equal to a break point falls into the lower level.  This is consistent
    with the paper's "tail" definition (at or below the mean).

    若 *breaks* 不足 3 個斷點（資料量過少或變異太小導致 head_tail_breaks
    提早終止），則以最後一個斷點重複填充至 3 個，確保不會因 IndexError 崩潰。
    If *breaks* has fewer than 3 elements (too few data points or low
    variance), pad with the last break value to prevent IndexError.
    """
    # ── 防禦：確保至少 3 個斷點 ──
    if len(breaks) == 0:
        # 無斷點 → 全部歸為 NC
        return ['NC'] * len(values)
    while len(breaks) < 3:
        breaks = list(breaks) + [breaks[-1]]

    labels = []
    for v in values:
        if v <= breaks[0]:
            labels.append('NC')
        elif v <= breaks[1]:
            labels.append('C-III')
        elif v <= breaks[2]:
            labels.append('C-II')
        else:
            labels.append('C-I')
    return labels


def compute_epidemic_risk(g, town_data, d, daytime, number_of_loops=5000,
                           progress_callback=None):
    """Core EpiRank algorithm — Stages 2 & 3 of the pipeline.

    This function takes the raw commuting network (Stage 1 output) and
    computes the stationary epidemic risk vector through two phases:

    Phase A — Matrix normalisation (Stage 2)
    ─────────────────────────────────────────
    The raw OD matrix is column-normalised into two stochastic matrices,
    each capturing a different direction of disease transmission:

        W   = col-normalise(OD)     → 「回家」backward / push direction
        W^T = col-normalise(OD^T)   → 「上班」forward  / pull direction

    Why column-normalisation?  Imagine township j has 3 commuting sources:
    A sends 600, B sends 300, C sends 100 commuters.  Column-normalising
    converts these to transition probabilities: A→j = 0.6, B→j = 0.3,
    C→j = 0.1.  This means 60% of j's infection risk from incoming
    commuters comes from A.  The absolute commuter count is factored out,
    leaving the *relative connectivity structure* — exactly what matters
    for epidemic spread patterns.

    Phase B — Iterative convergence (Stage 3)
    ──────────────────────────────────────────
    Starting from a uniform distribution (every township equally risky),
    the algorithm repeatedly applies:

        ER(t+1) = (1-d) · (1/N)  +  d · [daytime · W^T · ER(t)
                                       + (1-daytime) · W  · ER(t)]

    Intuition for each term:
    ┌──────────────────────┬─────────────────────────────────────────┐
    │ (1-d) · (1/N)        │ Teleportation / external factor:        │
    │                      │ with probability (1-d), a pathogen      │
    │                      │ arrives from an external source (e.g.   │
    │                      │ international travel, random contact)   │
    │                      │ regardless of the commuting network.    │
    │                      │ This prevents isolated islands from     │
    │                      │ having zero risk.                       │
    ├──────────────────────┼─────────────────────────────────────────┤
    │ d · daytime · W^T·ER │ Forward (morning) commute contribution: │
    │                      │ commuters ARRIVE at their workplaces,   │
    │                      │ carrying risk FROM their home townships.│
    │                      │ W^T propagates risk in the direction    │
    │                      │ home → work.  High-risk townships that  │
    │                      │ SEND many workers raise the risk of     │
    │                      │ the destination (pull effect).          │
    ├──────────────────────┼─────────────────────────────────────────┤
    │ d·(1-daytime)· W·ER  │ Backward (evening) commute contribution:│
    │                      │ commuters RETURN to their residences,   │
    │                      │ carrying risk FROM their workplaces.    │
    │                      │ W propagates risk in the direction      │
    │                      │ work → home.  High-risk workplaces push │
    │                      │ disease back to the bedroom suburbs     │
    │                      │ (push effect).                          │
    └──────────────────────┴─────────────────────────────────────────┘

    The elegance: by adjusting a single parameter ``daytime``, the model
    smoothly interpolates between three epidemiologically distinct regimes:
      - daytime=0.0: purely backward (evening return; disease spreads to
                     residential areas — like a flu brought home to family)
      - daytime=0.5: bidirectional (realistic; both directions contribute
                     equally — the paper's recommended default)
      - daytime=1.0: purely forward (morning arrival; disease accumulates
                     at workplaces — like a nosocomial outbreak)

    ╔══════════════════════════════════════════════════════════════════╗
    ║  為什麼 EpiRank 保證收斂？ — 數學證明                            ║
    ║  Why is EpiRank guaranteed to converge? — Mathematical proof   ║
    ╚══════════════════════════════════════════════════════════════════╝

    令 P = α·W' + (1−α)·W，迭代公式可改寫為：
    Let P = α·W' + (1−α)·W, then the iteration becomes:

        ER(t) = M · ER(t−1),  where  M = (1−d)·E + d·P
                                      E = (1/N)·1·1^T  (uniform rank-1 matrix)

    收斂性由以下四個環環相扣的性質保證：
    Convergence is guaranteed by four interlocking properties:

    ┌─────────────────────────────────────────────────────────────────┐
    │ 性質 1 (Property 1): P 是 column-stochastic 矩陣               │
    │ P is column-stochastic                                         │
    │                                                                │
    │   W 和 W' 各自 column-stochastic（每行總和 = 1）。               │
    │   P 是兩者的凸組合（α 和 1−α 非負、和為 1），                     │
    │   因此 P 本身也是 column-stochastic。                            │
    │                                                                │
    │   W and W' are each column-stochastic (each column sums to 1). │
    │   P is their convex combination (α + (1-α) = 1), so P is also  │
    │   column-stochastic.                                           │
    ├─────────────────────────────────────────────────────────────────┤
    │ 性質 2 (Property 2): Google Matrix M 是嚴格正矩陣               │
    │ The Google Matrix M is strictly positive                       │
    │                                                                │
    │   E 的每個元素都是 (1−d)/N > 0（因為 0 < d < 1）。               │
    │   即使 P 中有零元素，加上 (1−d)·E 後，M 的每一個元素              │
    │   都嚴格大於零。M 同時也是 column-stochastic                     │
    │  （兩個 column-stochastic 矩陣的凸組合）。                       │
    │                                                                │
    │   Every entry of E equals (1-d)/N > 0 (since 0 < d < 1).      │
    │   Even if P contains zeros, adding (1-d)·E makes every entry   │
    │   of M strictly positive.  M is also column-stochastic (convex │
    │   combination of two column-stochastic matrices).              │
    ├─────────────────────────────────────────────────────────────────┤
    │ 性質 3 (Property 3): Perron-Frobenius 定理直接適用              │
    │ The Perron-Frobenius theorem applies directly                  │
    │                                                                │
    │   M 是正的 column-stochastic 矩陣 → 不可約且非週期               │
    │     • 唯一最大特徵值 λ₁ = 1                                     │
    │     • 所有其他特徵值 |λᵢ| < 1（嚴格小於 1）                      │
    │     • 對應 λ₁ = 1 的特徵向量即為唯一穩態分布 ER*                 │
    │                                                                │
    │   M is a positive column-stochastic matrix → irreducible and   │
    │   aperiodic.                                                   │
    │     • Unique dominant eigenvalue: lambda_1 = 1                  │
    │     • All other eigenvalues: |lambda_i| < 1 (strictly)         │
    │     • The eigenvector for lambda_1 = 1 is the unique           │
    │       stationary distribution ER*                              │
    │                                                                │
    │   Power iteration from ANY initial vector converges to ER*.    │
    ├─────────────────────────────────────────────────────────────────┤
    │ 性質 4 (Property 4): 收斂速率是幾何級數，由 d 控制              │
    │ Convergence rate is geometric, controlled by d                 │
    │                                                                │
    │   第二大特徵值滿足 |λ₂| <= d，因此：                             │
    │   The second-largest eigenvalue satisfies |lambda_2| <= d:     │
    │                                                                │
    │       ||ER(t) - ER*|| <= d^t · ||ER(0) - ER*||                 │
    │                                                                │
    │       d = 0.85 → 50 次迭代後誤差衰減至 ~3e-4                    │
    │       d = 0.95 → 50 次迭代後誤差衰減至 ~0.077                   │
    │                  100 次迭代後 ~0.006                             │
    │       d = 0.95 (50 iters): error ~ 0.077                      │
    │       d = 0.95 (100 iters): error ~ 0.006                     │
    │                                                                │
    │   d 越大 → 網路結構影響力越大，但收斂較慢                        │
    │   d 越小 → 收斂越快，但結果退化為均勻分布                        │
    │   Larger d → more network influence, slower convergence        │
    │   Smaller d → faster convergence, but result → uniform         │
    └─────────────────────────────────────────────────────────────────┘

    直覺總結 / Intuitive summary:
    (1−d)·e 這一項是整個收斂的關鍵。它扮演 PageRank 中「隨機跳躍」
    (teleportation) 的角色——保證每個節點在每一步都有非零機率被「造訪」，
    從而消除了死胡同 (dangling nodes) 和週期性 (periodicity) 兩個阻礙
    收斂的因素。只要 d < 1，矩陣 M 就是嚴格正矩陣，Perron-Frobenius
    定理便給出無條件的收斂保證。

    The (1-d)·e term is the key to convergence.  It plays the role of
    PageRank's "teleportation" — ensuring every node has a non-zero
    probability of being "visited" at each step, thereby eliminating
    both dangling nodes and periodicity — the two obstacles to
    convergence.  As long as d < 1, M is a strictly positive matrix
    and the Perron-Frobenius theorem provides an unconditional
    convergence guarantee.

    Typically converges within 50–200 iterations for d=0.95.

    Parameters
    ----------
    g : nx.DiGraph
        The commuting network built by ``build_commuting_network()``.
    town_data : dict
        Township metadata (not modified; used only for key lookups).
    d : float
        Damping factor (0, 1).  Higher values (e.g. 0.95) give more
        weight to the network structure; lower values (e.g. 0.50) make
        the result more uniform.  Analogous to PageRank's alpha.
    daytime : float
        Forward/backward balance (0, 1).  See table above.
    number_of_loops : int
        Maximum iterations.  Convergence typically occurs well before
        this limit.
    progress_callback : callable or None
        Optional ``callback(current_iter, max_iter)`` for progress bars.

    Returns
    -------
    epidemic_risk : np.ndarray, shape (N, 1)
        The converged EpiRank vector.  Sums to 1.0.  Each element ER[i]
        represents the *relative* epidemic risk of the i-th node (ordered
        by ``list(g.nodes())``).
    iterations : int
        Actual number of iterations performed before convergence.
    CN_C : np.ndarray, shape (N, N)
        The raw (un-normalised) OD commuting count matrix.  Returned for
        use in downstream analyses (e.g. commuter flow statistics).
    """
    Ncount = g.order()

    # ══════════════════════════════════════════════════════════════
    # Phase A: Matrix Construction & Normalisation (Stage 2)
    # ══════════════════════════════════════════════════════════════

    # ── Step A1: Extract the raw OD matrix from the graph (Eq. 1) ──
    #
    # CN_C[i,j] = number of commuters from node i to node j.
    #
    # The row/column ordering follows list(g.nodes()) — the insertion
    # order of nodes in the DiGraph, preserved by nx.to_numpy_array().
    # This ordering is CRITICAL: it must be used consistently everywhere
    # that maps between matrix indices and node identifiers.
    #
    # Example (3 townships):
    #                To: A    B    C
    #   From: A  [[ 800  200   50 ]    ← 800 locals, 200 commute A→B
    #         B   [ 150  600  100 ]
    #         C   [  30   80  500 ]]
    #
    CN_C = nx.to_numpy_array(g, weight=KEY_COMMUTER_TYPE1)
    CN_T = CN_C.copy()

    # ── Step A2: Column-normalise OD → W (backward matrix, Eq. 2) ──
    #
    # W[i,j] = OD[i,j] / Σ_k OD[k,j]
    #
    # Each column of W sums to 1.0, forming a column-stochastic matrix.
    # Interpretation: W[i,j] is the probability that a commuter arriving
    # at township j came from township i.
    #
    # W models the *backward* (evening) commuting pattern.  When we
    # multiply W · ER, township j's risk "flows back" to all the home
    # townships i in proportion to their commuter share.  This captures
    # the scenario: a worker gets infected at workplace j, then carries
    # the pathogen home to residence i.
    #
    # Continuing the example:
    #   Column B sums to 200 + 600 + 80 = 880
    #   W[:,B] = [200/880, 600/880, 80/880] = [0.227, 0.682, 0.091]
    #   → 68.2% of B's "backward risk" stays local, 22.7% flows to A
    #
    csum = CN_T.sum(axis=0)
    CN = np.zeros((Ncount, Ncount))
    for i in range(CN_T.shape[0]):
        for j in range(CN_T.shape[1]):
            s = float(csum[j])
            if s > 0:
                CN[i, j] = float(CN_T[i, j]) / s

    # ── Step A3: Column-normalise OD^T → W^T (forward matrix, Eq. 3) ──
    #
    # First transpose the raw OD matrix, then column-normalise.
    # CNt = col-normalise(OD^T)
    #
    # Interpretation: CNt[j,i] is the probability that a commuter
    # leaving home township i goes to workplace township j.
    #
    # CNt models the *forward* (morning) commuting pattern.  When we
    # multiply CNt · ER, township i's risk "flows forward" to all the
    # workplaces j that its residents commute to.  This captures the
    # scenario: infected residents of township i carry the pathogen
    # to their various workplaces.
    #
    # Key asymmetry: W and CNt generally produce DIFFERENT risk rankings.
    # A large bedroom suburb (many outgoing commuters) will have high
    # risk under the backward model (W) because workers bring disease
    # home.  A city-centre business district (many incoming commuters)
    # will have high risk under the forward model (CNt) because it
    # attracts infected commuters from many sources.
    #
    ODt = CN_T.T
    osum = ODt.sum(axis=0)
    CNt = np.zeros((Ncount, Ncount))
    for i in range(ODt.shape[0]):
        for j in range(ODt.shape[1]):
            s = float(osum[j])
            if s > 0:
                CNt[i, j] = float(ODt[i, j]) / s

    # ══════════════════════════════════════════════════════════════
    # Phase B: Iterative Power Iteration (Stage 3)
    # ══════════════════════════════════════════════════════════════

    # ── Step B1: Initialise ER vector and external factor (Eq. 5) ──
    #
    # Start with uniform distribution: every township has equal risk
    # (1/N).  The external factor vector is also uniform — representing
    # "background noise" of infection from sources outside the commuting
    # network (e.g. international travellers, random community contact).
    #
    # The initial distribution doesn't affect the final result (the
    # stationary vector is unique), but uniform is a natural choice and
    # converges faster than a random starting point.
    #
    other_factors = np.ones((Ncount, 1)) / float(Ncount)
    epidemic_risk = np.ones((Ncount, 1)) / float(Ncount)

    # ── Step B2: Iterate until convergence (Eq. 4) ──
    #
    #   ER(t+1) = (1-d) · exFac  +  d · [ daytime   · CNt · ER(t)    ← forward
    #                                    + (1-daytime) · CN · ER(t) ]  ← backward
    #
    # This is a power iteration on a modified stochastic matrix.  At
    # each step, the new risk of township i is a weighted blend of:
    #   1. A uniform "teleportation" baseline              [(1-d)/N]
    #   2. Risk propagated forward through morning commute [d·daytime·CNt·ER]
    #   3. Risk propagated backward through evening return [d·(1-daytime)·CN·ER]
    #
    # 收斂保證 / Convergence guarantee:
    #   CN (W) 和 CNt (W') 皆為 column-stochastic，其凸組合 P 亦然。
    #   加入 (1-d)·e 後，迭代矩陣 M = (1-d)·E + d·P 為嚴格正矩陣。
    #   由 Perron-Frobenius 定理，M 有唯一最大特徵值 λ₁=1，
    #   所有 |λᵢ|<1，故 power iteration 必定收斂至唯一穩態分布。
    #   收斂速率：||ER(t)-ER*|| ≤ d^t · ||ER(0)-ER*||
    #   （完整數學證明見上方 docstring「為什麼 EpiRank 保證收斂」一節）
    #
    #   CN (W) and CNt (W') are column-stochastic; their convex combo
    #   P is too.  Adding (1-d)·e makes M = (1-d)·E + d·P strictly
    #   positive.  By Perron-Frobenius: unique λ₁=1, all |λᵢ|<1,
    #   so power iteration converges to the unique stationary vector.
    #   Rate: ||ER(t)-ER*|| <= d^t · ||ER(0)-ER*||
    #   (Full proof: see docstring section "Why is EpiRank guaranteed
    #    to converge?" above.)
    #
    # The ``@`` operator performs matrix multiplication (NumPy >= 1.10),
    # replacing the old np.asmatrix() * pattern from the original ERA.py.
    #
    # 收斂判準：元素最大變化量 < 1e-12，通常 50–200 次迭代即收斂 (d=0.95)
    # Convergence criterion: max element-wise change < 1e-12.
    # Typical convergence: 50–200 iterations for d=0.95.
    #
    iterations = 0
    for i in range(number_of_loops):
        old_er = epidemic_risk.copy()
        epidemic_risk = ((1.0 - d) * other_factors +
                         d * (daytime * (CNt @ epidemic_risk) +
                              (1.0 - daytime) * (CN @ epidemic_risk)))
        iterations = i + 1

        if np.allclose(epidemic_risk, old_er, atol=1e-12):
            break

        if progress_callback and (i % 50 == 0 or i == number_of_loops - 1):
            progress_callback(i + 1, number_of_loops)

    # ER 向量收斂後加總為 1.0 — 每個元素代表該鄉鎮佔全台灣疫情風險的
    # 「份額」。值越大，表示該鄉鎮在通勤網絡中越處於疫情傳播的樞紐位置。
    #
    # After convergence, ER sums to 1.0 — each element represents the
    # township's "share" of the total epidemic risk across all of Taiwan.
    # Higher values indicate that the township sits at a critical hub
    # in the commuting network for disease transmission.

    return epidemic_risk, iterations, CN_C


# ============================================================
# Worker Thread — runs EpiRank computation off the GUI thread
# ============================================================

class ComputeWorker(QThread):
    """Background thread that loads data, builds the commuting network,
    computes EpiRank (and PageRank / HITS for comparison), and emits
    the complete results dict on ``finished_ok``.

    The worker temporarily ``os.chdir()`` into ``data_dir`` so that all
    data-loading functions can use relative paths.  A ``finally`` block
    ensures the original working directory is always restored.
    """
    progress = Signal(int, int)       # current, max
    log_message = Signal(str)
    finished_ok = Signal(dict)        # results dict
    finished_err = Signal(str)

    def __init__(self, data_dir, d, daytime, max_loops):
        super().__init__()
        self.data_dir = data_dir
        self.d = d
        self.daytime = daytime
        self.max_loops = max_loops

    def run(self):
        old_cwd = os.getcwd()
        try:
            os.chdir(self.data_dir)

            # ── Stage 1: Load all data files ──
            town_data = {}
            g = nx.DiGraph()

            self.log_message.emit("Loading basic town data (bs.xlsx)...")
            build_basic_table_of_towns(town_data)
            self.log_message.emit(f"  Loaded {len(town_data)} towns.")

            self.log_message.emit("Loading Flu reported cases (Flu.xlsx)...")
            build_flu_reported_cases(town_data)

            self.log_message.emit("Loading Enterovirus reported cases (ev.xlsx)...")
            build_ev_reported_cases(town_data)

            self.log_message.emit("Loading SARS reported cases (SARS.xlsx)...")
            build_sars_reported_cases(town_data)

            self.log_message.emit("Building commuting network (cn.xlsx)... (this may take a while)")
            build_commuting_network(g, town_data)
            self.log_message.emit(f"  Network: {g.number_of_nodes()} nodes, {g.number_of_edges()} edges")

            # ── Stage 2: Compute EpiRank (main run) ──
            self.log_message.emit(f"Computing EpiRank (d={self.d}, daytime={self.daytime}, max_loops={self.max_loops})...")
            epidemic_risk, iterations, CN_C = compute_epidemic_risk(
                g, town_data, self.d, self.daytime, self.max_loops,
                progress_callback=lambda cur, mx: self.progress.emit(cur, mx)
            )
            self.log_message.emit(f"  Converged after {iterations} iterations.")

            # ── Stage 3: Compute EpiRank for 3 canonical daytime values ──
            # These are needed for Figures 6, 7, and 8 which compare
            # daytime = 0.0 (backward), 0.5 (bidirectional), 1.0 (forward).
            self.log_message.emit("Computing EpiRank for daytime=0.0, 0.5, 1.0 (for Figure 6)...")
            fig6_daytimes = [0.0, 0.5, 1.0]
            fig6_data = {}
            for dt in fig6_daytimes:
                if abs(dt - self.daytime) < 1e-9:
                    # Reuse the already-computed result
                    fig6_data[dt] = {
                        'epidemic_risk': epidemic_risk,
                        'iterations': iterations,
                    }
                else:
                    er_dt, it_dt, _ = compute_epidemic_risk(
                        g, town_data, self.d, dt, self.max_loops)
                    fig6_data[dt] = {
                        'epidemic_risk': er_dt,
                        'iterations': it_dt,
                    }
                    self.log_message.emit(
                        f"  daytime={dt}: converged after {fig6_data[dt]['iterations']} iterations.")

            # ── Stage 4: Compute alternative network indices for comparison ──
            # Paper Section III.B compares EpiRank against PageRank and HITS.
            self.log_message.emit("Computing PageRank, HITS...")
            page_rank = nx.pagerank(g, alpha=self.d)
            hub_rank, authority_rank = nx.hits(g, max_iter=1000)

            # ── Stage 5: Build result dictionaries keyed by seq_no ──
            # CRITICAL: list(g.nodes()) must be used (not sorted) to match
            # the row/column order used by nx.to_numpy_array() in
            # compute_epidemic_risk().
            nodes = list(g.nodes())
            ER_rank = {seq_no: float(epidemic_risk[i, 0]) for i, seq_no in enumerate(nodes)}
            pop_rank = {seq_no: town_data[g.nodes[seq_no][KEY_DB_ID]][KEY_POPULATION] for seq_no in nodes}
            flu_case_rank = {seq_no: town_data[g.nodes[seq_no][KEY_DB_ID]].get(KEY_FLU_TOTAL_CASES, 0) for seq_no in nodes}
            ev_case_rank = {seq_no: town_data[g.nodes[seq_no][KEY_DB_ID]].get(KEY_EV_AVERAGE_CASES, 0) for seq_no in nodes}
            sars_case_rank = {seq_no: town_data[g.nodes[seq_no][KEY_DB_ID]].get(KEY_SARS_TOTAL_CASES, 0) for seq_no in nodes}

            # ── Stage 6: Compute summary statistics ──
            er_values = np.array(list(ER_rank.values()))
            ER_tot = float(er_values.sum())
            ER_avg = float(er_values.mean())
            ER_std = float(er_values.std())

            # ── Stage 7: Compute correlations (Paper Table 2) ──
            correlations = {}
            for name, rank_dict in [('PageRank', page_rank), ('HITS Hub', hub_rank),
                                      ('HITS Authority', authority_rank), ('EpiRank', ER_rank),
                                      ('Population', pop_rank)]:
                corr_row = {}
                for disease, case_dict in [('Flu', flu_case_rank), ('EV', ev_case_rank), ('SARS', sars_case_rank)]:
                    common_keys = set(rank_dict.keys()) & set(case_dict.keys())
                    if len(common_keys) > 2:
                        sub1 = {k: rank_dict[k] for k in common_keys}
                        sub2 = {k: case_dict[k] for k in common_keys}
                        pr, pp = get_pearson_cor(sub1, sub2)
                        sr, sp = get_spearman_cor(sub1, sub2)
                        kr, kp = get_kendalltau_cor(sub1, sub2)
                        corr_row[disease] = {
                            'Pearson': (pr, pp), 'Spearman': (sr, sp), 'Kendall': (kr, kp)
                        }
                correlations[name] = corr_row

            # ── Stage 8: Build sorted result table for Tab 0 display ──
            ER_sorted = sorted_map(ER_rank)
            table_data = []
            for seq_no, er_val in ER_sorted:
                db_ID = g.nodes[seq_no][KEY_DB_ID]
                td = town_data[db_ID]
                table_data.append({
                    'seq_no': seq_no,
                    'post_code': g.nodes[seq_no][KEY_POST_CODE],
                    'db_ID': db_ID,
                    'county': td[KEY_COUNTY],
                    'town': td[KEY_TOWN],
                    'population': td[KEY_POPULATION],
                    'area': round(td[KEY_AREA], 2),
                    'density': round(td[KEY_DENSITY], 2),
                    'ERV': round(er_val, 8),
                    'ERP': round(100.0 * er_val / ER_tot, 4) if ER_tot > 0 else 0,
                    'C_local': td[KEY_LOCAL_COMMUTER_TYPE1],
                    'C_out': td[KEY_OUT_COMMUTER_TYPE1],
                    'C_in': td[KEY_IN_COMMUTER_TYPE1],
                    'page_rank': round(page_rank.get(seq_no, 0), 8),
                    'hub': round(hub_rank.get(seq_no, 0), 8),
                    'authority': round(authority_rank.get(seq_no, 0), 8),
                    'flu_cases': td.get(KEY_FLU_TOTAL_CASES, 0),
                    'ev_cases': td.get(KEY_EV_AVERAGE_CASES, 0),
                    'sars_cases': td.get(KEY_SARS_TOTAL_CASES, 0),
                    'pos_x': td[KEY_POS_XY][0],
                    'pos_y': td[KEY_POS_XY][1],
                })

            # Network positions for drawing (TWD97 TM2 coordinates in metres)
            npos = {seq_no: town_data[g.nodes[seq_no][KEY_DB_ID]][KEY_POS_XY] for seq_no in nodes}

            results = {
                'g': g,
                'town_data': town_data,
                'epidemic_risk': epidemic_risk,
                'table_data': table_data,
                'ER_rank': ER_rank,
                'ER_tot': ER_tot, 'ER_avg': ER_avg, 'ER_std': ER_std,
                'correlations': correlations,
                'page_rank': page_rank,
                'hub_rank': hub_rank,
                'authority_rank': authority_rank,
                'flu_case_rank': flu_case_rank,
                'ev_case_rank': ev_case_rank,
                'sars_case_rank': sars_case_rank,
                'pop_rank': pop_rank,
                'iterations': iterations,
                'CN_C': CN_C,
                'npos': npos,
                'd': self.d,
                'daytime': self.daytime,
                'fig6_data': fig6_data,
            }
            self.finished_ok.emit(results)

        except Exception as e:
            import traceback
            self.finished_err.emit(f"{e}\n\n{traceback.format_exc()}")
        finally:
            os.chdir(old_cwd)


# ============================================================
# Sensitivity Analysis Worker (Figure 11)
# ============================================================

class SensitivityWorker(QThread):
    """Compute EpiRank for a grid of (daytime, d) parameter pairs
    and calculate Pearson/Spearman correlations against disease data."""
    progress = Signal(int, int)       # current, total
    log_message = Signal(str)
    finished_ok = Signal(dict)        # sensitivity results
    finished_err = Signal(str)

    def __init__(self, data_dir, max_loops):
        super().__init__()
        self.data_dir = data_dir
        self.max_loops = max_loops

    def run(self):
        """Sweep over daytime × d parameter grid and compute correlations.

        Paper Figure 11: "Sensitivity analysis of correlation between
        EpiRank and actual disease data for different values of daytime
        parameter and damping factor."

        Grid: daytime ∈ [0.0, 0.05, ..., 1.0] (21 values)
              d       ∈ [0.05, 0.10, ..., 1.0] (20 values)
        Total: 420 parameter combinations.
        """
        old_cwd = os.getcwd()
        try:
            os.chdir(self.data_dir)

            self.log_message.emit("=== Sensitivity Analysis ===")
            self.log_message.emit("Loading data files...")

            town_data = {}
            g = nx.DiGraph()
            build_basic_table_of_towns(town_data)
            build_flu_reported_cases(town_data)
            build_ev_reported_cases(town_data)
            build_commuting_network(g, town_data)

            # CRITICAL: list(g.nodes()) preserves insertion order to match
            # nx.to_numpy_array() index mapping used inside compute_epidemic_risk().
            nodes = list(g.nodes())
            flu_case_rank = {s: town_data[g.nodes[s][KEY_DB_ID]].get(KEY_FLU_TOTAL_CASES, 0) for s in nodes}
            ev_case_rank  = {s: town_data[g.nodes[s][KEY_DB_ID]].get(KEY_EV_AVERAGE_CASES, 0) for s in nodes}

            # Parameter grid
            daytime_values = np.round(np.arange(0.0, 1.05, 0.05), 2)   # 21 values
            d_values       = np.round(np.arange(0.05, 1.025, 0.05), 2) # 20 values
            total = len(daytime_values) * len(d_values)  # 420

            flu_pearson  = np.zeros((len(daytime_values), len(d_values)))
            flu_spearman = np.zeros((len(daytime_values), len(d_values)))
            ev_pearson   = np.zeros((len(daytime_values), len(d_values)))
            ev_spearman  = np.zeros((len(daytime_values), len(d_values)))

            count = 0
            for i, dt in enumerate(daytime_values):
                for j, dv in enumerate(d_values):
                    er, iters, _ = compute_epidemic_risk(g, town_data, float(dv), float(dt), self.max_loops)
                    er_vals = np.array(er).flatten()
                    er_rank = {nodes[k]: er_vals[k] for k in range(len(nodes))}

                    fp_r, _ = get_pearson_cor(er_rank, flu_case_rank)
                    fs_r, _ = get_spearman_cor(er_rank, flu_case_rank)
                    ep_r, _ = get_pearson_cor(er_rank, ev_case_rank)
                    es_r, _ = get_spearman_cor(er_rank, ev_case_rank)

                    flu_pearson[i, j]  = fp_r
                    flu_spearman[i, j] = fs_r
                    ev_pearson[i, j]   = ep_r
                    ev_spearman[i, j]  = es_r

                    count += 1
                    self.progress.emit(count, total)

                self.log_message.emit(
                    f"  daytime={dt:.2f} completed ({count}/{total})")

            self.log_message.emit(f"\nSensitivity analysis complete ({total} parameter combinations).")
            self.finished_ok.emit({
                'flu_pearson': flu_pearson,
                'flu_spearman': flu_spearman,
                'ev_pearson': ev_pearson,
                'ev_spearman': ev_spearman,
                'daytime_values': daytime_values,
                'd_values': d_values,
            })

        except Exception as e:
            import traceback
            self.finished_err.emit(f"{e}\n\n{traceback.format_exc()}")
        finally:
            os.chdir(old_cwd)


# ============================================================
# Main Window
# ============================================================

class EpiRankMainWindow(QMainWindow):
    """Main application window.

    Manages the parameter controls, 14 result tabs, background computation
    workers, and file export.  ``self.results`` holds the full output dict
    from ``ComputeWorker``; ``self.sensitivity_results`` holds the output
    from ``SensitivityWorker``.  Both are ``None`` until their respective
    computations complete.
    """

    def __init__(self):
        super().__init__()
        self.setWindowTitle("EpiRank - Epidemic Risk Analysis System")
        self.setMinimumSize(1200, 800)

        self.results = None               # populated by ComputeWorker.finished_ok
        self.sensitivity_results = None   # populated by SensitivityWorker.finished_ok
        self.worker = None                # reference to the active QThread
        self.data_dir = os.path.dirname(os.path.abspath(__file__))

        self._build_menu()
        self._build_ui()
        self._update_status("Ready. Set parameters and click 'Run EpiRank'.")

    # ---- Menu ----
    def _build_menu(self):
        menubar = self.menuBar()

        file_menu = menubar.addMenu("File (&F)")
        act_set_dir = QAction("Set Data Directory...", self)
        act_set_dir.triggered.connect(self._choose_data_dir)
        file_menu.addAction(act_set_dir)

        act_export = QAction("Export Results to Excel...", self)
        act_export.triggered.connect(self._export_excel)
        file_menu.addAction(act_export)

        act_save_fig = QAction("Save Current Chart...", self)
        act_save_fig.triggered.connect(self._save_figure)
        file_menu.addAction(act_save_fig)

        file_menu.addSeparator()
        act_quit = QAction("Quit (&Q)", self)
        act_quit.triggered.connect(self.close)
        file_menu.addAction(act_quit)

    # ---- UI ----
    def _build_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        main_layout = QVBoxLayout(central)

        # Top: Parameters + Controls
        top_widget = QWidget()
        top_layout = QHBoxLayout(top_widget)
        top_layout.setContentsMargins(0, 0, 0, 0)

        # Parameter group
        param_group = QGroupBox("EpiRank Parameters")
        param_form = QFormLayout()

        self.spin_d = QDoubleSpinBox()
        self.spin_d.setRange(0.01, 0.99)
        self.spin_d.setSingleStep(0.01)
        self.spin_d.setValue(0.95)
        self.spin_d.setDecimals(2)
        self.spin_d.setToolTip("Damping factor (d): probability of following commuting links.\n"
                                "Higher d means commuting patterns dominate the risk score.")
        param_form.addRow("Damping Factor (d):", self.spin_d)

        self.spin_daytime = QDoubleSpinBox()
        self.spin_daytime.setRange(0.01, 0.99)
        self.spin_daytime.setSingleStep(0.01)
        self.spin_daytime.setValue(0.50)
        self.spin_daytime.setDecimals(2)
        self.spin_daytime.setToolTip("Daytime weight: proportion of day spent at work destination.\n"
                                      "0.5 means equal time at origin and destination.")
        param_form.addRow("Daytime Weight:", self.spin_daytime)

        self.spin_loops = QSpinBox()
        self.spin_loops.setRange(100, 50000)
        self.spin_loops.setSingleStep(100)
        self.spin_loops.setValue(5000)
        self.spin_loops.setToolTip("Maximum number of iterations for EpiRank convergence.")
        param_form.addRow("Max Iterations:", self.spin_loops)

        param_group.setLayout(param_form)
        top_layout.addWidget(param_group)

        # Data directory display
        dir_group = QGroupBox("Data Directory")
        dir_layout = QVBoxLayout()
        self.lbl_data_dir = QLabel(self.data_dir)
        self.lbl_data_dir.setWordWrap(True)
        dir_layout.addWidget(self.lbl_data_dir)
        btn_dir = QPushButton("Change...")
        btn_dir.clicked.connect(self._choose_data_dir)
        dir_layout.addWidget(btn_dir)
        dir_group.setLayout(dir_layout)
        top_layout.addWidget(dir_group)

        # Run button
        ctrl_group = QGroupBox("Control")
        ctrl_layout = QVBoxLayout()
        self.btn_run = QPushButton("Run EpiRank")
        self.btn_run.setStyleSheet("font-size: 16px; font-weight: bold; padding: 10px;")
        self.btn_run.clicked.connect(self._run_computation)
        ctrl_layout.addWidget(self.btn_run)

        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        ctrl_layout.addWidget(self.progress_bar)

        self.btn_sensitivity = QPushButton("Sensitivity Analysis")
        self.btn_sensitivity.setStyleSheet("font-size: 13px; padding: 6px;")
        self.btn_sensitivity.setEnabled(False)
        self.btn_sensitivity.setToolTip(
            "Run EpiRank for all daytime × damping factor combinations\n"
            "to generate sensitivity heatmaps (Figure 11).\n"
            "Requires 420 computations — may take several minutes.")
        self.btn_sensitivity.clicked.connect(self._run_sensitivity)
        ctrl_layout.addWidget(self.btn_sensitivity)

        ctrl_group.setLayout(ctrl_layout)
        top_layout.addWidget(ctrl_group)

        main_layout.addWidget(top_widget)

        # ── Bottom: Tab widget containing all result views ──
        # Layout: 14 tabs (0–13) mapping to paper figures/tables.
        # See file-level docstring for the complete tab-to-figure mapping.
        self.tabs = QTabWidget()

        # Tab 0: Results Table — ranked EpiRank scores with region filter
        table_widget = QWidget()
        table_layout = QVBoxLayout(table_widget)
        filter_bar = QHBoxLayout()
        filter_bar.addWidget(QLabel("篩選區域:"))
        self.combo_region = QComboBox()
        self.combo_region.addItems([
            "全台灣 353 鄉鎮市",
            "大台北 gTaipei（48 鄉鎮市）",
        ])
        self.combo_region.currentIndexChanged.connect(self._populate_table)
        filter_bar.addWidget(self.combo_region)
        filter_bar.addStretch()
        table_layout.addLayout(filter_bar)
        self.table = QTableWidget()
        self.table.setAlternatingRowColors(True)
        table_layout.addWidget(self.table)
        self.tabs.addTab(table_widget, "Results Table")

        # Tab 1: Network Map
        self.fig_network = Figure(figsize=(5, 10), dpi=100, facecolor='white')
        self.canvas_network = FigureCanvas(self.fig_network)
        self.toolbar_network = NavigationToolbar(self.canvas_network, self)
        net_widget = QWidget()
        net_layout = QVBoxLayout(net_widget)
        net_layout.addWidget(self.toolbar_network)
        net_layout.addWidget(self.canvas_network)
        self.tabs.addTab(net_widget, "Network Map")

        # Tab 2: Core Classification (Table 1 from paper)
        self.table1_widget = QTableWidget()
        self.table1_widget.setAlternatingRowColors(True)
        self.tabs.addTab(self.table1_widget, "Core Classification")

        # Tab 3: Correlation Table (Table 2 from paper)
        self.corr_table = QTableWidget()
        self.corr_table.setAlternatingRowColors(True)
        self.tabs.addTab(self.corr_table, "Correlations")

        # Tab 4: Data Analysis (Figure 2 from paper)
        self.fig_analysis = Figure(figsize=(14, 18), dpi=100, facecolor='white')
        self.canvas_analysis = FigureCanvas(self.fig_analysis)
        self.toolbar_analysis = NavigationToolbar(self.canvas_analysis, self)
        analysis_widget = QWidget()
        analysis_layout = QVBoxLayout(analysis_widget)
        analysis_layout.addWidget(self.toolbar_analysis)
        analysis_layout.addWidget(self.canvas_analysis)
        self.tabs.addTab(analysis_widget, "Commuter Flow")

        # Tab 5: Frequency Distributions (Figure 3 from paper)
        self.fig_disease = Figure(figsize=(12, 10), dpi=100, facecolor='white')
        self.canvas_disease = FigureCanvas(self.fig_disease)
        self.toolbar_disease = NavigationToolbar(self.canvas_disease, self)
        disease_widget = QWidget()
        disease_layout = QVBoxLayout(disease_widget)
        disease_layout.addWidget(self.toolbar_disease)
        disease_layout.addWidget(self.canvas_disease)
        self.tabs.addTab(disease_widget, "Frequency Distributions")

        # Tab 6: Frequency Distribution (Figure 6 from paper)
        self.fig_epirank_dist = Figure(figsize=(14, 8), dpi=100, facecolor='white')
        self.canvas_epirank_dist = FigureCanvas(self.fig_epirank_dist)
        self.toolbar_epirank_dist = NavigationToolbar(self.canvas_epirank_dist, self)
        epirank_dist_widget = QWidget()
        epirank_dist_layout = QVBoxLayout(epirank_dist_widget)
        epirank_dist_layout.addWidget(self.toolbar_epirank_dist)
        epirank_dist_layout.addWidget(self.canvas_epirank_dist)
        self.tabs.addTab(epirank_dist_widget, "Frequency Distribution")

        # Tab 7: EpiRank vs Disease (Figure 9 from paper)
        self.fig_epirank_vs = Figure(figsize=(12, 6), dpi=100, facecolor='white')
        self.canvas_epirank_vs = FigureCanvas(self.fig_epirank_vs)
        self.toolbar_epirank_vs = NavigationToolbar(self.canvas_epirank_vs, self)
        epirank_vs_widget = QWidget()
        epirank_vs_layout = QVBoxLayout(epirank_vs_widget)
        epirank_vs_layout.addWidget(self.toolbar_epirank_vs)
        epirank_vs_layout.addWidget(self.canvas_epirank_vs)
        self.tabs.addTab(epirank_vs_widget, "EpiRank vs Disease")

        # Tab 8: Index Comparison (Figure 10 from paper)
        self.fig_index_comp = Figure(figsize=(16, 5), dpi=100, facecolor='white')
        self.canvas_index_comp = FigureCanvas(self.fig_index_comp)
        self.toolbar_index_comp = NavigationToolbar(self.canvas_index_comp, self)
        index_comp_widget = QWidget()
        index_comp_layout = QVBoxLayout(index_comp_widget)
        index_comp_layout.addWidget(self.toolbar_index_comp)
        index_comp_layout.addWidget(self.canvas_index_comp)
        self.tabs.addTab(index_comp_widget, "Index Comparison")

        # Tab 9: Disease Map (Figure 4 from paper)
        self.fig_disease_map = Figure(figsize=(12, 6), dpi=100, facecolor='white')
        self.canvas_disease_map = FigureCanvas(self.fig_disease_map)
        self.toolbar_disease_map = NavigationToolbar(self.canvas_disease_map, self)
        disease_map_widget = QWidget()
        disease_map_layout = QVBoxLayout(disease_map_widget)
        disease_map_layout.addWidget(self.toolbar_disease_map)
        disease_map_layout.addWidget(self.canvas_disease_map)
        self.tabs.addTab(disease_map_widget, "Disease Map")

        # Tab 10: EpiRank Map (Figure 7 from paper)
        self.fig_epirank_map = Figure(figsize=(14, 10), dpi=100, facecolor='white')
        self.canvas_epirank_map = FigureCanvas(self.fig_epirank_map)
        self.toolbar_epirank_map = NavigationToolbar(self.canvas_epirank_map, self)
        epirank_map_widget = QWidget()
        epirank_map_layout = QVBoxLayout(epirank_map_widget)
        epirank_map_layout.addWidget(self.toolbar_epirank_map)
        epirank_map_layout.addWidget(self.canvas_epirank_map)
        self.tabs.addTab(epirank_map_widget, "EpiRank Map")

        # Tab 11: EpiRank vs Disease Map (Figure 8 from paper)
        self.fig_overlay_map = Figure(figsize=(12, 6), dpi=100, facecolor='white')
        self.canvas_overlay_map = FigureCanvas(self.fig_overlay_map)
        self.toolbar_overlay_map = NavigationToolbar(self.canvas_overlay_map, self)
        overlay_map_widget = QWidget()
        overlay_map_layout = QVBoxLayout(overlay_map_widget)
        overlay_map_layout.addWidget(self.toolbar_overlay_map)
        overlay_map_layout.addWidget(self.canvas_overlay_map)
        self.tabs.addTab(overlay_map_widget, "EpiRank vs Disease Map")

        # Tab 12: Log
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setFont(QFont("Courier", 11))
        self.tabs.addTab(self.log_text, "Log")

        # Tab 13: Sensitivity Analysis (Figure 11 from paper) — populated by Sensitivity button
        self.fig_sensitivity = Figure(figsize=(12, 10), dpi=100, facecolor='white')
        self.canvas_sensitivity = FigureCanvas(self.fig_sensitivity)
        self.toolbar_sensitivity = NavigationToolbar(self.canvas_sensitivity, self)
        sensitivity_widget = QWidget()
        sensitivity_layout = QVBoxLayout(sensitivity_widget)
        sensitivity_layout.addWidget(self.toolbar_sensitivity)
        sensitivity_layout.addWidget(self.canvas_sensitivity)
        self.tabs.addTab(sensitivity_widget, "Sensitivity Analysis")

        main_layout.addWidget(self.tabs, stretch=1)

        # Status bar
        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)

    # ---- Actions ----
    def _choose_data_dir(self):
        d = QFileDialog.getExistingDirectory(self, "Select Data Directory", self.data_dir)
        if d:
            self.data_dir = d
            self.lbl_data_dir.setText(d)

    def _update_status(self, msg):
        self.status_bar.showMessage(msg)

    def _log(self, msg):
        self.log_text.append(msg)

    def _run_computation(self):
        # Check required files
        required = ['bs.xlsx', 'Flu.xlsx', 'ev.xlsx', 'SARS.xlsx', 'cn.xlsx']
        missing = [f for f in required if not os.path.isfile(os.path.join(self.data_dir, f))]
        if missing:
            QMessageBox.warning(self, "Missing Data Files",
                                f"The following files are missing in {self.data_dir}:\n\n" +
                                "\n".join(missing))
            return

        self.btn_run.setEnabled(False)
        self.progress_bar.setValue(0)
        self.log_text.clear()
        self._log(f"Data directory: {self.data_dir}")
        self._log(f"Parameters: d={self.spin_d.value()}, daytime={self.spin_daytime.value()}, max_loops={self.spin_loops.value()}")

        self.worker = ComputeWorker(
            self.data_dir,
            self.spin_d.value(),
            self.spin_daytime.value(),
            self.spin_loops.value()
        )
        self.worker.progress.connect(self._on_progress)
        self.worker.log_message.connect(self._log)
        self.worker.finished_ok.connect(self._on_finished_ok)
        self.worker.finished_err.connect(self._on_finished_err)
        self.worker.start()
        self._update_status("Computing...")

    def _on_progress(self, cur, mx):
        if mx > 0:
            self.progress_bar.setValue(int(100 * cur / mx))

    def _on_finished_ok(self, results):
        self.results = results
        self.btn_run.setEnabled(True)
        self.progress_bar.setValue(100)
        self._log(f"\nDone! Converged in {results['iterations']} iterations.")
        self._log(f"ER Total={results['ER_tot']:.6f}, Mean={results['ER_avg']:.6f}, StdDev={results['ER_std']:.6f}")
        self._update_status(f"Completed. {len(results['table_data'])} towns analyzed, {results['iterations']} iterations.")

        # ── 每個分頁獨立 try/except，確保單一分頁失敗不影響其餘 ──
        # Each tab is wrapped individually so one failure does not block
        # the remaining tabs from being populated.
        _tab_tasks = [
            ("Results Table",          self._populate_table),
            ("Network Graph",          self._draw_network),
            ("Data Analysis",          self._draw_data_analysis),
            ("Disease Analysis",       self._draw_disease_analysis),
            ("EpiRank Distribution",   self._draw_epirank_distribution),
            ("EpiRank vs Disease",     self._draw_epirank_vs_disease),
            ("Index Comparison",       self._draw_index_comparison),
            ("Disease Map",            self._draw_disease_map),
            ("EpiRank Map",            self._draw_epirank_map),
            ("Overlay Map",            self._draw_overlay_map),
            ("Detail Table",           self._populate_table1),
            ("Correlations",           self._populate_correlations),
        ]
        for tab_name, func in _tab_tasks:
            try:
                func()
            except Exception as exc:
                self._log(f"WARNING: Failed to populate tab '{tab_name}': {exc}")

        # Enable Sensitivity Analysis button now that data is loaded
        self.btn_sensitivity.setEnabled(True)

        # Auto-save output files (same behavior as old ERA.py)
        self._auto_save_results()

        self.tabs.setCurrentIndex(0)

    # ---- 安全關閉視窗 ----
    def closeEvent(self, event):
        """確保背景 worker 結束後再關閉視窗，避免 RuntimeError。

        Wait for any running background worker to finish before closing,
        preventing 'wrapped C++ object has been deleted' RuntimeError
        when a QThread emits a signal to a destroyed widget.
        """
        for worker_attr in ('worker', 'sensitivity_worker'):
            w = getattr(self, worker_attr, None)
            if w is not None and w.isRunning():
                self._log(f"Waiting for {worker_attr} to finish...")
                w.quit()
                w.wait(5000)  # 最多等 5 秒 / wait up to 5 seconds
        event.accept()

    def _on_finished_err(self, err_msg):
        self.btn_run.setEnabled(True)
        self.progress_bar.setValue(0)
        self._log(f"\nERROR:\n{err_msg}")
        self._update_status("Error during computation.")
        QMessageBox.critical(self, "Computation Error", str(err_msg)[:500])

    # ---- Populate Results Table (Tab 0) ----
    def _populate_table(self):
        """Fill the Results Table (Tab 0) with ranked EpiRank scores.

        Row background colour uses a continuous mean ± σ scheme (5 bands)
        rather than the discrete head/tail breaks used elsewhere.  This
        gives a finer visual gradient in the table view:

            ERV >= mean + 2σ   →  Red         (very high risk)
            ERV >= mean + 1σ   →  Orange      (high risk)
            ERV >= mean        →  Yellow      (medium risk)
            ERV >= mean − 1σ   →  Light green (low risk)
            ERV <  mean − 1σ   →  Gray        (very low risk)

        The thresholds are always computed on the full 353-town statistics,
        even when the table is filtered to gTaipei (48 towns), so colours
        remain consistent across views.
        """
        if self.results is None:
            return

        all_data = self.results['table_data']

        # Filter by region selection
        if self.combo_region.currentIndex() == 1:
            # 大台北 gTaipei（48 鄉鎮市）
            data = [item for item in all_data if item['db_ID'] in GTAIPEI_DB_IDS]
        else:
            # 全台灣 353 鄉鎮市
            data = all_data

        headers = ['Rank', 'County', 'Town', 'ERV', 'ERP (%)',
                    'Population', 'Area', 'Density',
                    'C.local', 'C.out', 'C.in',
                    'PageRank', 'Hub', 'Authority',
                    'Flu Cases', 'EV Cases', 'SARS Cases']

        self.table.setRowCount(len(data))
        self.table.setColumnCount(len(headers))
        self.table.setHorizontalHeaderLabels(headers)

        ER_avg = self.results['ER_avg']
        ER_std = self.results['ER_std']

        for row, item in enumerate(data):
            erv = item['ERV']
            # Color by mean ± σ bands (NOT head/tail breaks; see docstring)
            if erv >= ER_avg + 2 * ER_std:
                bg = QColor(255, 100, 100)   # Red - very high
            elif erv >= ER_avg + ER_std:
                bg = QColor(255, 180, 100)   # Orange - high
            elif erv >= ER_avg:
                bg = QColor(255, 255, 150)   # Yellow - medium
            elif erv >= ER_avg - ER_std:
                bg = QColor(200, 230, 200)   # Light green - low
            else:
                bg = QColor(220, 220, 220)   # Gray - very low

            values = [
                row + 1, item['county'], item['town'],
                f"{erv:.8f}", f"{item['ERP']:.4f}%",
                f"{item['population']:.0f}", item['area'], item['density'],
                item['C_local'], item['C_out'], item['C_in'],
                f"{item['page_rank']:.8f}", f"{item['hub']:.8f}", f"{item['authority']:.8f}",
                item['flu_cases'], item['ev_cases'], item['sars_cases']
            ]
            for col, val in enumerate(values):
                cell = QTableWidgetItem(str(val))
                cell.setBackground(bg)
                if col >= 3:
                    cell.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.table.setItem(row, col, cell)

        self.table.resizeColumnsToContents()
        self.table.horizontalHeader().setStretchLastSection(True)

    # ---- Draw Network Map (Tab 1) ----
    def _draw_network(self):
        """Visualise the commuting network on a geographic map.

        Nodes are positioned using TWD97 TM2 coordinates (metres); node
        size reflects relative EpiRank score and colour encodes the
        mean ± σ risk level.  This map is also auto-saved as a PNG file.
        """
        self.fig_network.clear()
        ax = self.fig_network.add_subplot(111)

        g = self.results['g']
        npos = self.results['npos']  # TWD97 TM2 coordinates (meters)
        ER_rank = self.results['ER_rank']
        ER_avg = self.results['ER_avg']
        ER_std = self.results['ER_std']

        dg = nx.Graph(g)

        # Node colors and sizes by risk level
        node_list = list(dg.nodes())
        ncolor = []
        nsize = []
        for seq_no in node_list:
            er = ER_rank.get(seq_no, 0)
            nsize.append(max(er * 10, 0.002) ** 2 * 50000)
            if er >= ER_avg + 2 * ER_std:
                ncolor.append('red')
            elif er >= ER_avg + ER_std:
                ncolor.append('darkorange')
            elif er >= ER_avg:
                ncolor.append('gold')
            elif er >= ER_avg - ER_std:
                ncolor.append('lightgreen')
            else:
                ncolor.append('lightgray')

        # Normalize sizes to reasonable range
        max_size = max(nsize) if nsize else 1
        nsize = [s / max_size * 300 + 5 for s in nsize]

        # Draw edges first (underneath), then nodes
        nx.draw_networkx_edges(dg, pos=npos, width=0.1, alpha=0.12,
                                edge_color='steelblue', ax=ax)
        nx.draw_networkx_nodes(dg, pos=npos, nodelist=node_list,
                                node_size=nsize, node_color=ncolor,
                                linewidths=0.3, edgecolors='gray', ax=ax)

        ax.set_title(f"台灣鄉鎮市通勤網絡 (Taiwan Commuting Network)\n"
                      f"d={self.results['d']}, daytime={self.results['daytime']}",
                      fontsize=10)
        ax.set_aspect('equal')   # Preserve correct geographic proportions
        ax.axis('off')
        self.fig_network.tight_layout()
        self.canvas_network.draw()

    # ---- Draw Data Analysis (Figure 2 — Tab 4) ----
    def _draw_data_analysis(self):
        """Reproduce the 7 subplots of Figure 2 (Commuter Flow Data).

        Paper Section II.A: "Statistical properties of the commuting data."
        Layout is a 4-row × 3-column grid matching the paper:

          col 0    col 1    col 2
        ┌────────────────┬────────┐
        │                │  (b)   │  row 0
        │      (a)       ├────────┤
        │   density map  │  (c)   │  row 1
        │  3 rows × 2col ├────────┤
        │                │  (d)   │  row 2
        ├────────┬───────┼────────┤
        │  (e)   │  (f)  │  (g)   │  row 3
        └────────┴───────┴────────┘
        """
        from matplotlib.gridspec import GridSpec

        self.fig_analysis.clear()
        gs = GridSpec(4, 3, figure=self.fig_analysis,
                      hspace=0.35, wspace=0.35,
                      left=0.06, right=0.97, top=0.97, bottom=0.04)

        g = self.results['g']
        town_data = self.results['town_data']
        nodes = list(g.nodes())

        # Pre-compute per-node data
        db_ids    = [g.nodes[n][KEY_DB_ID] for n in nodes]
        positions = [town_data[db][KEY_POS_XY] for db in db_ids]
        xs        = [p[0] for p in positions]
        ys        = [p[1] for p in positions]
        norm_dens = [town_data[db][KEY_NORMALIZED_DENSITY] for db in db_ids]

        # Local commuter flow per node (self-loop edge weight)
        local_flows_arr = np.array(
            [g[n][n][KEY_COMMUTER_TYPE1] if g.has_edge(n, n) else 0
             for n in nodes], dtype=float)

        # Total in/out flows (includes local commuters)
        total_in  = np.array([town_data[db][KEY_IN_COMMUTER_TYPE1]
                              for db in db_ids], dtype=float)
        total_out = np.array([town_data[db][KEY_OUT_COMMUTER_TYPE1]
                              for db in db_ids], dtype=float)

        # Inter-township flows only (exclude local / self-loop)
        inter_in  = total_in  - local_flows_arr
        inter_out = total_out - local_flows_arr

        # Degrees excluding self-loops
        in_degs  = np.array([g.in_degree(n)  - (1 if g.has_edge(n, n) else 0)
                             for n in nodes], dtype=float)
        out_degs = np.array([g.out_degree(n) - (1 if g.has_edge(n, n) else 0)
                             for n in nodes], dtype=float)

        # ── (a) Township density map — spans rows 0-2, cols 0-1 ──
        ax_a = self.fig_analysis.add_subplot(gs[0:3, 0:2])
        # Classify: urbanized (norm_dens > 0.10), regular (0.01~0.10), rural (<0.01)
        urban_idx   = [i for i, nd in enumerate(norm_dens) if nd > 0.10]
        regular_idx = [i for i, nd in enumerate(norm_dens) if 0.01 < nd <= 0.10]
        rural_idx   = [i for i, nd in enumerate(norm_dens) if nd <= 0.01]

        ax_a.scatter([xs[i] for i in rural_idx],   [ys[i] for i in rural_idx],
                     s=4, c='green', alpha=0.6, label='rural')
        ax_a.scatter([xs[i] for i in regular_idx],  [ys[i] for i in regular_idx],
                     s=10, c='royalblue', alpha=0.7, label='regular')
        ax_a.scatter([xs[i] for i in urban_idx],    [ys[i] for i in urban_idx],
                     s=22, c='purple', alpha=0.8, label='urbanized')
        ax_a.set_aspect('equal')
        ax_a.set_title('(a)', fontsize=11)
        ax_a.legend(fontsize=8, loc='upper left', markerscale=1.5)
        ax_a.axis('off')

        # ── (b) In-degree vs Out-degree (excluding self-loops) — row 0, col 2 ──
        ax_b = self.fig_analysis.add_subplot(gs[0, 2])
        ax_b.scatter(in_degs, out_degs, s=8, c='steelblue', alpha=0.6)
        lim = (max(in_degs.max(), out_degs.max()) + 5) if len(in_degs) > 0 else 10
        ax_b.plot([0, lim], [0, lim], 'gray', linewidth=0.8, alpha=0.6)
        ax_b.set_xlim(0, lim)
        ax_b.set_ylim(0, lim)
        ax_b.set_xlabel('in degree', fontsize=8)
        ax_b.set_ylabel('out degree', fontsize=8)
        ax_b.set_title('(b)', fontsize=9)
        ax_b.tick_params(labelsize=6)

        # ── (c) Weighted In-flow vs Out-flow (inter-township only) — row 1, col 2 ──
        ax_c = self.fig_analysis.add_subplot(gs[1, 2])
        ax_c.scatter(inter_in / 1e4, inter_out / 1e4, s=8, c='steelblue', alpha=0.6)
        flow_lim = (max(inter_in.max(), inter_out.max()) / 1e4 * 1.1) if len(inter_in) > 0 else 1.0
        ax_c.plot([0, flow_lim], [0, flow_lim], 'gray', linewidth=0.8, alpha=0.6)
        ax_c.set_xlim(0, flow_lim)
        ax_c.set_ylim(0, flow_lim)
        ax_c.set_xlabel(r'in flow ($\times 10^4$)', fontsize=8)
        ax_c.set_ylabel(r'out flow ($\times 10^4$)', fontsize=8)
        ax_c.set_title('(c)', fontsize=9)
        ax_c.tick_params(labelsize=6)

        # ── (d) Log ratio of inter-township in/out — row 2, col 2 ──
        ax_d = self.fig_analysis.add_subplot(gs[2, 2])
        log_ratios = []
        for i in range(len(nodes)):
            if inter_out[i] > 0 and inter_in[i] > 0:
                log_ratios.append(np.log10(inter_in[i] / inter_out[i]))
            else:
                log_ratios.append(0.0)
        log_ratios_sorted = sorted(log_ratios, reverse=True)
        n_pull = sum(1 for v in log_ratios_sorted if v > 0)
        # Paper style: single blue filled area + vertical line at zero-crossing
        ax_d.bar(range(len(log_ratios_sorted)), log_ratios_sorted,
                 width=1.0, color='steelblue', alpha=0.8)
        ax_d.axvline(x=n_pull, color='blue', linewidth=1.2)
        ax_d.axhline(y=0, color='black', linewidth=0.5)
        ax_d.set_xlabel('township rank', fontsize=8)
        ax_d.set_ylabel('log10(in / out)', fontsize=8)
        ax_d.set_title('(d)', fontsize=9)
        ax_d.tick_params(labelsize=6)

        # ── Pre-compute edge distances for (e) and (f) ──
        edge_dists = []  # (distance_km, commuter_count)
        for u, v, edata in g.edges(data=True):
            if u == v:
                continue  # skip self-loops (local commuters)
            ux, uy = town_data[g.nodes[u][KEY_DB_ID]][KEY_POS_XY]
            vx, vy = town_data[g.nodes[v][KEY_DB_ID]][KEY_POS_XY]
            dist_km = np.sqrt((ux - vx)**2 + (uy - vy)**2) / 1000.0
            commuters = edata[KEY_COMMUTER_TYPE1]
            edge_dists.append((dist_km, commuters))

        # ── (e) Commuting distance distribution — row 3, col 0 ──
        ax_e = self.fig_analysis.add_subplot(gs[3, 0])
        if edge_dists:
            dist_arr = np.array([d for d, _ in edge_dists])
            weight_arr = np.array([w for _, w in edge_dists])
            total_commuters = weight_arr.sum()
            bins = np.arange(0, dist_arr.max() + 1, 1)  # 1 km bins
            ax_e.hist(dist_arr, bins=bins, weights=weight_arr / 1e4,
                      color='steelblue', edgecolor='white', linewidth=0.3, alpha=0.8)
            ax_e.set_xlabel('commuting distance (km)', fontsize=8)
            ax_e.set_ylabel(r'no. commuters ($\times 10^4$)', fontsize=8)
            ax_e.set_title('(e)', fontsize=9)
            ax_e.tick_params(labelsize=6)

        # ── (f) Cumulative commuting distance — row 3, col 1 ──
        ax_f = self.fig_analysis.add_subplot(gs[3, 1])
        if edge_dists:
            sorted_edges = sorted(edge_dists, key=lambda x: x[0])
            cum_commuters = 0
            cum_x, cum_y = [0], [0]
            for d_km, w in sorted_edges:
                cum_commuters += w
                cum_x.append(d_km)
                cum_y.append(100.0 * cum_commuters / total_commuters)
            ax_f.plot(cum_x, cum_y, 'steelblue', linewidth=1.5)
            ax_f.set_xlim(0, max(cum_x))
            ax_f.set_ylim(0, 105)
            # Add reference lines at 80% and 90%
            for pct in [80, 90]:
                ax_f.axhline(y=pct, color='gray', linestyle=':', linewidth=0.8)
                for cx, cy in zip(cum_x, cum_y):
                    if cy >= pct:
                        ax_f.axvline(x=cx, color='black', linestyle=':', linewidth=0.8)
                        ax_f.annotate(f'{pct}% -> {cx:.0f}km', xy=(cx, pct),
                                      fontsize=7, color='red',
                                      xytext=(cx + 1, pct - 5))
                        break
            ax_f.set_xlabel('commuting distance (km)', fontsize=8)
            ax_f.set_ylabel('cumulative %', fontsize=8)
            ax_f.set_title('(f)', fontsize=9)
            ax_f.tick_params(labelsize=6)

        # ── (g) Local flow distribution — row 3, col 2 ──
        # Histogram of local (self-loop) commuter flow per township.
        # Paper reports ~84% of commuters work within their own township.
        ax_g = self.fig_analysis.add_subplot(gs[3, 2])
        if local_flows_arr.max() > 0:
            bins_g = np.linspace(0, local_flows_arr.max() * 1.05, 11)  # 10 bins
            ax_g.hist(local_flows_arr, bins=bins_g, color='#d2691e',
                      edgecolor='white', linewidth=0.3, alpha=0.8)
            ax_g.set_xlabel('local flow', fontsize=8)
            ax_g.set_ylabel('no. townships', fontsize=8)
            ax_g.set_title('(g)', fontsize=9)
            ax_g.tick_params(labelsize=6)

        self.canvas_analysis.draw()

    # ---- Draw Disease Analysis (Figure 3 — Tab 5) ----
    def _draw_disease_analysis(self):
        """Reproduce the 4 subplots of Figure 3 (Disease Frequency & In/Out Ratio).

        Paper Section II.B: "Epidemic data and head/tail breaks."
        Layout is a 2-row × 2-column grid:

          col 0                col 1
        +--------------------+--------------------+
        | (a) freq dist flu  | (b) freq dist EV   |  row 0
        +--------------------+--------------------+
        | (c) in/out flu     | (d) in/out EV      |  row 1
        +--------------------+--------------------+
        """
        from matplotlib.gridspec import GridSpec

        self.fig_disease.clear()
        gs = GridSpec(2, 2, figure=self.fig_disease,
                      hspace=0.38, wspace=0.30,
                      left=0.08, right=0.96, top=0.94, bottom=0.08)

        g = self.results['g']
        town_data = self.results['town_data']
        nodes = list(g.nodes())
        db_ids = [g.nodes[n][KEY_DB_ID] for n in nodes]

        # ── Collect disease case data per township ──
        flu_cases = np.array([town_data[db].get(KEY_FLU_TOTAL_CASES, 0)
                              for db in db_ids], dtype=float)
        ev_cases  = np.array([town_data[db].get(KEY_EV_AVERAGE_CASES, 0.0)
                              for db in db_ids], dtype=float)

        # ── Compute inter-township log ratios (same as Data Analysis) ──
        local_flows_arr = np.array(
            [g[n][n][KEY_COMMUTER_TYPE1] if g.has_edge(n, n) else 0
             for n in nodes], dtype=float)
        total_in  = np.array([town_data[db][KEY_IN_COMMUTER_TYPE1]
                              for db in db_ids], dtype=float)
        total_out = np.array([town_data[db][KEY_OUT_COMMUTER_TYPE1]
                              for db in db_ids], dtype=float)
        inter_in  = total_in  - local_flows_arr
        inter_out = total_out - local_flows_arr

        log_ratios = np.zeros(len(nodes))
        for i in range(len(nodes)):
            if inter_out[i] > 0 and inter_in[i] > 0:
                log_ratios[i] = np.log10(inter_in[i] / inter_out[i])

        # ── Head/tail breaks: 3 breaks → 4 groups ──
        flu_breaks = head_tail_breaks(flu_cases, 3)
        ev_breaks  = head_tail_breaks(ev_cases, 3)

        # ── Assign epidemic level per township ──
        flu_levels = classify_by_breaks(flu_cases, flu_breaks)
        ev_levels  = classify_by_breaks(ev_cases, ev_breaks)

        # ── Helper: draw frequency distribution histogram ──
        def draw_freq_dist(ax, case_values, breaks, levels, title,
                           xlabel='epidemic risk'):
            """
            Plot (a) or (b): histogram of case counts, bars colored by
            head/tail level.  Each integer case-count is one bar whose
            colour is determined by the break-point range it belongs to.
            """
            vmin = max(0, case_values.min())
            vmax = case_values.max()
            # Integer bins for integer data, finer bins for continuous
            if vmax <= 50 and np.allclose(case_values, case_values.astype(int)):
                bins = np.arange(vmin, vmax + 2, 1) - 0.5
            else:
                bins = np.linspace(vmin - 0.25, vmax + 0.25, 35)

            # Separate data by level and stack
            level_data = {lv: [] for lv in LEVEL_ORDER}
            for val, lv in zip(case_values, levels):
                level_data[lv].append(val)

            # Plot stacked histogram
            bottom = np.zeros(len(bins) - 1)
            for lv in LEVEL_ORDER:
                if not level_data[lv]:
                    continue
                counts, _ = np.histogram(level_data[lv], bins=bins)
                ax.bar(bins[:-1] + np.diff(bins) / 2, counts,
                       width=np.diff(bins) * 0.92,
                       bottom=bottom, color=LEVEL_COLORS[lv],
                       edgecolor='white', linewidth=0.3, label=lv)
                bottom += counts

            # Vertical dashed lines at break points (paper uses dashed)
            for bk in breaks:
                ax.axvline(x=bk, color='black', linestyle='--', linewidth=0.9)

            # Level labels at top, positioned between break lines (blue, matching paper)
            ylim_top = ax.get_ylim()[1]
            edges = [vmin] + list(breaks) + [vmax]
            for i, lv in enumerate(LEVEL_ORDER):
                if i < len(edges) - 1:
                    mid_x = (edges[i] + edges[i + 1]) / 2
                    ax.text(mid_x, ylim_top * 0.97, lv, ha='center',
                            va='top', fontsize=8, fontweight='bold',
                            color='blue')

            ax.set_xlabel(xlabel, fontsize=9)
            ax.set_ylabel('number of townships', fontsize=9)
            ax.set_title(title, fontsize=10)
            ax.tick_params(labelsize=7)

        # ── Helper: draw in/out ratio with epidemic level coloring ──
        def draw_inout_ratio(ax, log_ratios, levels, title):
            """
            Plot (c) or (d): stacked histogram of log10(in/out) colored
            by epidemic level, with push/pull labels.
            Paper uses 10 bins from -1.0 to 1.0 (bin width 0.2).
            """
            bins = np.linspace(-1.0, 1.0, 11)   # 10 bins, width 0.2

            # Separate by level
            level_data = {lv: [] for lv in LEVEL_ORDER}
            for lr, lv in zip(log_ratios, levels):
                level_data[lv].append(lr)

            # Stacked histogram
            bottom = np.zeros(len(bins) - 1)
            for lv in LEVEL_ORDER:
                if not level_data[lv]:
                    continue
                counts, _ = np.histogram(level_data[lv], bins=bins)
                ax.bar(bins[:-1] + np.diff(bins) / 2, counts,
                       width=np.diff(bins) * 0.92,
                       bottom=bottom, color=LEVEL_COLORS[lv],
                       edgecolor='white', linewidth=0.3, label=lv)
                bottom += counts

            # Push/pull vertical dashed line at x=0
            ax.axvline(x=0, color='black', linestyle='--', linewidth=1.0)

            # Push / Pull labels at upper corners (matching paper)
            ylim = ax.get_ylim()
            y_label = ylim[1] * 0.95 if ylim[1] > 0 else 1
            ax.text(-0.95, y_label, 'push', ha='left', va='top',
                    fontsize=9, fontweight='bold', color='#333333')
            ax.text(0.95, y_label, 'pull', ha='right', va='top',
                    fontsize=9, fontweight='bold', color='#333333')

            ax.set_xlim(-1.0, 1.0)
            ax.set_xlabel(r'$log_{10}$(in/out)', fontsize=9)
            ax.set_ylabel('number of townships', fontsize=9)
            ax.set_title(title, fontsize=10)
            ax.tick_params(labelsize=7)

        # ── Draw the 4 subplots ──
        ax_a = self.fig_disease.add_subplot(gs[0, 0])
        draw_freq_dist(ax_a, flu_cases, flu_breaks, flu_levels,
                       '(a) frequency distribution of flu cases')

        ax_b = self.fig_disease.add_subplot(gs[0, 1])
        draw_freq_dist(ax_b, ev_cases, ev_breaks, ev_levels,
                       '(b) frequency distribution of EV cases')

        ax_c = self.fig_disease.add_subplot(gs[1, 0])
        draw_inout_ratio(ax_c, log_ratios, flu_levels,
                         '(c) in/out ratio of flu cases')

        ax_d = self.fig_disease.add_subplot(gs[1, 1])
        draw_inout_ratio(ax_d, log_ratios, ev_levels,
                         '(d) in/out ratio of EV cases')

        self.canvas_disease.draw()

    # ---- Draw EpiRank Distribution (Figure 6 — Tab 6) ----
    def _draw_epirank_distribution(self):
        """Reproduce Figure 6: EpiRank frequency distributions and log in/out
        ratio distributions for 3 daytime settings (0.0, 0.5, 1.0).

        Paper Section III.A: "Figure 6 shows the distribution of EpiRank
        values and their corresponding in/out ratio distributions for
        three daytime parameter values."

        Each column corresponds to a different daytime value; bars are
        coloured by head/tail breaks core levels (NC, C-III, C-II, C-I).

          daytime=0.0         daytime=0.5         daytime=1.0
        +-----------------+-----------------+-----------------+
        | (a) freq dist   | (b) freq dist   | (c) freq dist   |  row 0
        +-----------------+-----------------+-----------------+
        | (d) log in/out  | (e) log in/out  | (f) log in/out  |  row 1
        +-----------------+-----------------+-----------------+
        """
        from matplotlib.gridspec import GridSpec

        self.fig_epirank_dist.clear()
        gs = GridSpec(2, 3, figure=self.fig_epirank_dist,
                      hspace=0.42, wspace=0.30,
                      left=0.07, right=0.97, top=0.92, bottom=0.08)

        g = self.results['g']
        town_data = self.results['town_data']
        nodes = list(g.nodes())
        db_ids = [g.nodes[n][KEY_DB_ID] for n in nodes]
        fig6_data = self.results['fig6_data']

        # Compute inter-township log ratios (same logic as Disease Analysis)
        local_flows_arr = np.array(
            [g[n][n][KEY_COMMUTER_TYPE1] if g.has_edge(n, n) else 0
             for n in nodes], dtype=float)
        total_in = np.array([town_data[db][KEY_IN_COMMUTER_TYPE1]
                             for db in db_ids], dtype=float)
        total_out = np.array([town_data[db][KEY_OUT_COMMUTER_TYPE1]
                              for db in db_ids], dtype=float)
        inter_in = total_in - local_flows_arr
        inter_out = total_out - local_flows_arr

        log_ratios = np.zeros(len(nodes))
        for i in range(len(nodes)):
            if inter_out[i] > 0 and inter_in[i] > 0:
                log_ratios[i] = np.log10(inter_in[i] / inter_out[i])

        # Sorted daytime values for columns
        daytimes = sorted(fig6_data.keys())
        subplot_labels_row1 = ['(a)', '(b)', '(c)']
        subplot_labels_row2 = ['(d)', '(e)', '(f)']

        for col_idx, dt in enumerate(daytimes):
            er_matrix = fig6_data[dt]['epidemic_risk']
            er_values = np.array([float(er_matrix[i, 0])
                                  for i in range(len(nodes))])

            # Head/tail breaks on EpiRank values
            er_breaks = head_tail_breaks(er_values, 3)
            er_levels = classify_by_breaks(er_values, er_breaks)

            # ── Row 0: Frequency distribution of EpiRank values ──
            ax_top = self.fig_epirank_dist.add_subplot(gs[0, col_idx])
            bins = np.linspace(0, max(er_values.max() * 1.05, 0.016), 30)

            level_data = {lv: [] for lv in LEVEL_ORDER}
            for val, lv in zip(er_values, er_levels):
                level_data[lv].append(val)

            bottom = np.zeros(len(bins) - 1)
            for lv in LEVEL_ORDER:
                if not level_data[lv]:
                    continue
                counts, _ = np.histogram(level_data[lv], bins=bins)
                ax_top.bar(bins[:-1] + np.diff(bins) / 2, counts,
                           width=np.diff(bins) * 0.92,
                           bottom=bottom, color=LEVEL_COLORS[lv],
                           edgecolor='white', linewidth=0.3, label=lv)
                bottom += counts

            # Vertical dashed lines at break points (paper uses dashed)
            for bk in er_breaks:
                ax_top.axvline(x=bk, color='black', linestyle='--', linewidth=0.9)

            # Level labels at top (abbreviated: NC, III, II, I; blue color)
            ylim_top = ax_top.get_ylim()[1]
            edges = [0] + list(er_breaks) + [er_values.max() * 1.1]
            abbrev_labels = ['NC', 'III', 'II', 'I']
            for i_lv, lv_abbrev in enumerate(abbrev_labels):
                if i_lv < len(edges) - 1:
                    mid_x = (edges[i_lv] + edges[i_lv + 1]) / 2
                    ax_top.text(mid_x, ylim_top * 0.97, lv_abbrev, ha='center',
                                va='top', fontsize=8, fontweight='bold',
                                color='blue')

            # "Core" annotation at upper right (matching paper)
            ax_top.annotate('Core', xy=(er_breaks[-1], ylim_top * 0.5),
                            xytext=(er_values.max() * 0.85, ylim_top * 0.7),
                            fontsize=9, fontweight='bold', color='#555555',
                            arrowprops=dict(arrowstyle='->', color='#555555',
                                            lw=1.2))

            ax_top.set_xlabel('EpiRank', fontsize=9)
            if col_idx == 0:
                ax_top.set_ylabel('number of townships', fontsize=9)
            ax_top.set_title(
                f'{subplot_labels_row1[col_idx]} daytime={dt:.1f}',
                fontsize=10)
            ax_top.tick_params(labelsize=7)

            # ── Row 1: Log in/out ratio colored by EpiRank level ──
            ax_bot = self.fig_epirank_dist.add_subplot(gs[1, col_idx])
            ratio_bins = np.linspace(-1.0, 1.0, 11)  # 10 bins, width 0.2

            ratio_level_data = {lv: [] for lv in LEVEL_ORDER}
            for lr, lv in zip(log_ratios, er_levels):
                ratio_level_data[lv].append(lr)

            bottom = np.zeros(len(ratio_bins) - 1)
            for lv in LEVEL_ORDER:
                if not ratio_level_data[lv]:
                    continue
                counts, _ = np.histogram(ratio_level_data[lv],
                                         bins=ratio_bins)
                ax_bot.bar(ratio_bins[:-1] + np.diff(ratio_bins) / 2,
                           counts,
                           width=np.diff(ratio_bins) * 0.92,
                           bottom=bottom, color=LEVEL_COLORS[lv],
                           edgecolor='white', linewidth=0.3, label=lv)
                bottom += counts

            ax_bot.axvline(x=0, color='black', linestyle='--', linewidth=1.0)

            ylim_b = ax_bot.get_ylim()
            y_label = ylim_b[1] * 0.95 if ylim_b[1] > 0 else 1
            ax_bot.text(-0.95, y_label, 'push', ha='left', va='top',
                        fontsize=9, fontweight='bold', color='#333333')
            ax_bot.text(0.95, y_label, 'pull', ha='right', va='top',
                        fontsize=9, fontweight='bold', color='#333333')

            ax_bot.set_xlim(-1.0, 1.0)
            ax_bot.set_xlabel(r'$log_{10}$(in/out)', fontsize=9)
            if col_idx == 0:
                ax_bot.set_ylabel('number of townships', fontsize=9)
            ax_bot.set_title(
                f'{subplot_labels_row2[col_idx]} daytime={dt:.1f}',
                fontsize=10)
            ax_bot.tick_params(labelsize=7)

        self.canvas_epirank_dist.draw()

    # ---- Draw EpiRank vs Disease (Figure 9 — Tab 7) ----
    def _draw_epirank_vs_disease(self):
        """Reproduce Figure 9: Cross-tabulation of actual disease core levels
        vs EpiRank-predicted core levels as stacked percentage bar charts.

        Paper Section III.C: "The cross-tabulation results provide insight
        into how well EpiRank's predicted core/non-core classification
        matches the observed epidemic severity."

        +-----------------------+-----------------------+
        | (a) flu case          | (b) EV case           |
        +-----------------------+-----------------------+
        X-axis: actual condition (core-I, core-II, core-III, non-core)
        Y-axis: percentage (0-100%)
        Stacked bars: predicted condition by EpiRank
        Numbers at top: actual case count in each group
        """
        from matplotlib.gridspec import GridSpec
        from matplotlib.patches import Patch

        self.fig_epirank_vs.clear()
        gs = GridSpec(1, 2, figure=self.fig_epirank_vs,
                      wspace=0.30,
                      left=0.08, right=0.95, top=0.88, bottom=0.12)

        g = self.results['g']
        town_data = self.results['town_data']
        nodes = list(g.nodes())
        db_ids = [g.nodes[n][KEY_DB_ID] for n in nodes]

        # EpiRank values (use daytime=0.5 if available, else current)
        fig6_data = self.results['fig6_data']
        dt_key = 0.5 if 0.5 in fig6_data else self.results['daytime']
        er_matrix = fig6_data[dt_key]['epidemic_risk']
        er_values = np.array([float(er_matrix[i, 0])
                              for i in range(len(nodes))])

        # Head/tail breaks on EpiRank → predicted core levels
        er_breaks = head_tail_breaks(er_values, 3)
        er_levels = classify_by_breaks(er_values, er_breaks)

        # Collect disease data
        flu_cases = np.array([town_data[db].get(KEY_FLU_TOTAL_CASES, 0)
                              for db in db_ids], dtype=float)
        ev_cases = np.array([town_data[db].get(KEY_EV_AVERAGE_CASES, 0.0)
                             for db in db_ids], dtype=float)

        # Head/tail breaks on actual disease cases → actual core levels
        flu_breaks = head_tail_breaks(flu_cases, 3)
        ev_breaks = head_tail_breaks(ev_cases, 3)
        flu_actual = classify_by_breaks(flu_cases, flu_breaks)
        ev_actual = classify_by_breaks(ev_cases, ev_breaks)

        # Reversed level order for x-axis: core-I first (matching paper)
        X_LEVELS = ['core-I', 'core-II', 'core-III', 'non-core']
        # Map our labels to x-axis labels
        LEVEL_TO_X = {'C-I': 'core-I', 'C-II': 'core-II',
                      'C-III': 'core-III', 'NC': 'non-core'}

        PREDICTED_COLORS = {
            'core-I':    '#cc2020',
            'core-II':   '#e07830',
            'core-III':  '#c8b840',
            'non-core':  '#3a8f3e',
        }

        def draw_comparison(ax, actual_levels, predicted_levels, title):
            """
            Draw a stacked percentage bar chart.
            actual_levels: list of actual core labels per township
            predicted_levels: list of EpiRank predicted labels per township
            """
            # Map to x-axis labels
            actual_x = [LEVEL_TO_X[a] for a in actual_levels]
            predicted_x = [LEVEL_TO_X[p] for p in predicted_levels]

            # For each actual group, count predicted distribution
            group_counts = {}
            group_totals = {}
            for a, p in zip(actual_x, predicted_x):
                if a not in group_counts:
                    group_counts[a] = {xl: 0 for xl in X_LEVELS}
                    group_totals[a] = 0
                group_counts[a][p] += 1
                group_totals[a] += 1

            # Build percentage arrays
            x_positions = np.arange(len(X_LEVELS))
            bar_width = 0.6

            bottom = np.zeros(len(X_LEVELS))
            for pred_lv in X_LEVELS:
                pcts = []
                for act_lv in X_LEVELS:
                    total = group_totals.get(act_lv, 0)
                    count = group_counts.get(act_lv, {}).get(pred_lv, 0)
                    pct = (100.0 * count / total) if total > 0 else 0
                    pcts.append(pct)

                ax.bar(x_positions, pcts, bar_width,
                       bottom=bottom,
                       color=PREDICTED_COLORS[pred_lv],
                       edgecolor='white', linewidth=0.5,
                       label=pred_lv)
                bottom += np.array(pcts)

            # Number annotations at top of each bar
            for i, act_lv in enumerate(X_LEVELS):
                total = group_totals.get(act_lv, 0)
                ax.text(i, 103, str(total), ha='center', va='bottom',
                        fontsize=9, fontweight='bold')

            ax.set_xticks(x_positions)
            ax.set_xticklabels(X_LEVELS, fontsize=9)
            ax.set_xlabel('actual condition', fontsize=10)
            ax.set_ylabel('percentage', fontsize=10)
            ax.set_ylim(0, 115)
            ax.set_yticks([0, 20, 40, 60, 80, 100])
            ax.set_yticklabels(['0 %', '20 %', '40 %', '60 %', '80 %', '100 %'],
                               fontsize=8)
            ax.set_title(title, fontsize=11)
            ax.tick_params(labelsize=8)

        # (a) Flu case
        ax_a = self.fig_epirank_vs.add_subplot(gs[0, 0])
        draw_comparison(ax_a, flu_actual, er_levels, '(a) flu case')

        # (b) EV case
        ax_b = self.fig_epirank_vs.add_subplot(gs[0, 1])
        draw_comparison(ax_b, ev_actual, er_levels, '(b) EV case')

        # Shared legend
        legend_patches = [
            Patch(facecolor=PREDICTED_COLORS[lv], label=lv)
            for lv in X_LEVELS
        ]
        self.fig_epirank_vs.legend(
            handles=legend_patches, loc='upper right',
            title='predicted condition', fontsize=8, title_fontsize=9,
            frameon=True, bbox_to_anchor=(0.98, 0.98))

        self.canvas_epirank_vs.draw()

    # ---- Populate Core Classification Table (Table 1 — Tab 2) ----
    def _populate_table1(self):
        """Reproduce Table 1: Numbers of core/non-core townships identified
        by EpiRank, PageRank, HITS-Hub, and HITS-Authority.

        Paper Section III.A: "Table 1 shows the numbers of identified
        core-I, core-II, core-III, and non-core townships."

        Note: Due to floating-point sensitivity at head/tail break
        boundaries (±10⁻¹⁸), counts may differ by ±1 from the paper's
        published values for isolated island townships (e.g. 綠島鄉, 蘭嶼鄉).

                         EpiRank  PageRank  HITS-Hub  HITS-Authority
        core-I              16       15        1          1
        core-II             31       20        2          3
        core-III            67       72       23         25
        non-core           239      246      327        324
        """
        if self.results is None:
            return

        nodes = list(self.results['g'].nodes())

        # Metrics to classify
        metrics = [
            ('EpiRank', self.results['ER_rank']),
            ('PageRank', self.results['page_rank']),
            ('HITS-Hub', self.results['hub_rank']),
            ('HITS-Authority', self.results['authority_rank']),
        ]

        row_labels = ['core-I', 'core-II', 'core-III', 'non-core']
        level_to_row = {'C-I': 0, 'C-II': 1, 'C-III': 2, 'NC': 3}
        row_colors = [
            QColor(255, 120, 120),   # core-I: red
            QColor(255, 190, 120),   # core-II: orange
            QColor(255, 255, 150),   # core-III: yellow
            QColor(180, 230, 180),   # non-core: green
        ]

        self.table1_widget.setRowCount(len(row_labels))
        self.table1_widget.setColumnCount(len(metrics) + 1)
        self.table1_widget.setHorizontalHeaderLabels(
            [''] + [m[0] for m in metrics])
        self.table1_widget.setVerticalHeaderLabels([])

        # Row labels in first column
        for r, label in enumerate(row_labels):
            item = QTableWidgetItem(label)
            item.setFont(QFont('Arial', 11, QFont.Bold))
            item.setBackground(row_colors[r])
            self.table1_widget.setItem(r, 0, item)

        for col_idx, (metric_name, metric_dict) in enumerate(metrics):
            values = np.array([metric_dict.get(n, 0) for n in nodes])
            breaks = head_tail_breaks(values, 3)
            levels = classify_by_breaks(values, breaks)

            # Count each level
            counts = {'C-I': 0, 'C-II': 0, 'C-III': 0, 'NC': 0}
            for lv in levels:
                counts[lv] += 1

            for lv, count in counts.items():
                r = level_to_row[lv]
                item = QTableWidgetItem(str(count))
                item.setTextAlignment(Qt.AlignCenter)
                item.setFont(QFont('Arial', 12))
                item.setBackground(row_colors[r])
                self.table1_widget.setItem(r, col_idx + 1, item)

        self.table1_widget.resizeColumnsToContents()
        self.table1_widget.horizontalHeader().setStretchLastSection(True)

        # Set minimum column width for readability
        for c in range(self.table1_widget.columnCount()):
            if self.table1_widget.columnWidth(c) < 120:
                self.table1_widget.setColumnWidth(c, 120)

    # ---- Draw Index Comparison (Figure 10 — Tab 8) ----
    def _draw_index_comparison(self):
        """Reproduce Figure 10: Distribution results from four network indices
        expressed as log₁₀(in/out) ratios.

        Paper Section III.B: "We compare EpiRank against PageRank, HITS-Hub,
        and HITS-Authority.  Figure 10 shows core (red) vs non-core (green)
        township distributions along the push-pull spectrum.  EpiRank
        uniquely identifies core townships in both push and pull regions."
        """
        from matplotlib.gridspec import GridSpec

        self.fig_index_comp.clear()
        gs = GridSpec(1, 4, figure=self.fig_index_comp,
                      wspace=0.30,
                      left=0.05, right=0.97, top=0.88, bottom=0.14)

        g = self.results['g']
        town_data = self.results['town_data']
        nodes = list(g.nodes())
        db_ids = [g.nodes[n][KEY_DB_ID] for n in nodes]

        # Compute log in/out ratios
        local_flows_arr = np.array(
            [g[n][n][KEY_COMMUTER_TYPE1] if g.has_edge(n, n) else 0
             for n in nodes], dtype=float)
        total_in = np.array([town_data[db][KEY_IN_COMMUTER_TYPE1]
                             for db in db_ids], dtype=float)
        total_out = np.array([town_data[db][KEY_OUT_COMMUTER_TYPE1]
                              for db in db_ids], dtype=float)
        inter_in = total_in - local_flows_arr
        inter_out = total_out - local_flows_arr

        log_ratios = np.zeros(len(nodes))
        for i in range(len(nodes)):
            if inter_out[i] > 0 and inter_in[i] > 0:
                log_ratios[i] = np.log10(inter_in[i] / inter_out[i])

        # Metrics to compare
        metrics = [
            ('(a) EpiRank', self.results['ER_rank']),
            ('(b) PageRank', self.results['page_rank']),
            ('(c) HITS-Hub', self.results['hub_rank']),
            ('(d) HITS-Authority', self.results['authority_rank']),
        ]

        COLOR_CORE = '#cc2020'      # red for core
        COLOR_NONCORE = '#3a8f3e'   # green for non-core

        for col_idx, (title, metric_dict) in enumerate(metrics):
            ax = self.fig_index_comp.add_subplot(gs[0, col_idx])

            values = np.array([metric_dict.get(n, 0) for n in nodes])
            breaks = head_tail_breaks(values, 3)
            levels = classify_by_breaks(values, breaks)

            # Binary: core vs non-core
            core_ratios = [lr for lr, lv in zip(log_ratios, levels)
                           if lv != 'NC']
            noncore_ratios = [lr for lr, lv in zip(log_ratios, levels)
                              if lv == 'NC']

            bins = np.linspace(-1.0, 1.0, 11)  # 10 bins, width 0.2

            # Stacked histogram: non-core (green) at bottom, core (red) on top
            nc_counts, _ = np.histogram(noncore_ratios, bins=bins)
            c_counts, _ = np.histogram(core_ratios, bins=bins)
            bar_centers = bins[:-1] + np.diff(bins) / 2
            bar_width = np.diff(bins) * 0.92

            ax.bar(bar_centers, nc_counts, width=bar_width,
                   color=COLOR_NONCORE, edgecolor='white', linewidth=0.3,
                   label='non-core')
            ax.bar(bar_centers, c_counts, width=bar_width,
                   bottom=nc_counts, color=COLOR_CORE,
                   edgecolor='white', linewidth=0.3,
                   label='core')

            ax.axvline(x=0, color='black', linestyle='--', linewidth=1.0)

            ax.set_xlim(-1.0, 1.0)
            ax.set_ylim(0, 80)
            ax.text(-0.95, 76, 'push', ha='left', va='top',
                    fontsize=9, fontweight='bold', color='#333333')
            ax.text(0.95, 76, 'pull', ha='right', va='top',
                    fontsize=9, fontweight='bold', color='#333333')

            ax.set_xlabel(r'$log_{10}$(in/out)', fontsize=9)
            if col_idx == 0:
                ax.set_ylabel('number of townships', fontsize=9)
            ax.set_title(title, fontsize=10)
            ax.tick_params(labelsize=7)

        # Shared legend
        from matplotlib.patches import Patch
        legend_handles = [
            Patch(facecolor=COLOR_CORE, label='core'),
            Patch(facecolor=COLOR_NONCORE, label='non-core'),
        ]
        self.fig_index_comp.legend(handles=legend_handles,
                                   loc='upper right', fontsize=9,
                                   frameon=True, bbox_to_anchor=(0.99, 0.99))

        self.canvas_index_comp.draw()

    # ---- Disease Map (Figure 4 — Tab 9) ----
    def _draw_disease_map(self):
        """Reproduce Figure 4: Spatial distributions of disease case severity.

        Paper Section II.B: "Figure 4 shows the spatial distributions of
        flu and EV case severity levels across Taiwan's 353 townships."

        Each subplot uses head/tail breaks on the disease case counts to
        classify townships into 4 levels (NC, C-III, C-II, C-I), plotted
        on TWD97 geographic coordinates.
        """
        self.fig_disease_map.clear()

        g = self.results['g']
        npos = self.results['npos']
        town_data = self.results['town_data']

        disease_info = [
            ('(a) flu case distribution', self.results['flu_case_rank']),
            ('(b) EV case distribution', self.results['ev_case_rank']),
        ]
        size_map = {'C-I': 120, 'C-II': 60, 'C-III': 30, 'NC': 10}
        legend_labels = {'NC': 'non-core', 'C-III': 'core-III',
                         'C-II': 'core-II', 'C-I': 'core-I'}

        for col, (title, case_dict) in enumerate(disease_info):
            ax = self.fig_disease_map.add_subplot(1, 2, col + 1)
            vals = np.array(list(case_dict.values()))
            keys = list(case_dict.keys())
            breaks = head_tail_breaks(vals)
            if len(breaks) < 3:
                ax.set_title(title + '\n(insufficient data for classification)')
                continue
            labels = classify_by_breaks(vals, breaks)

            # Draw each level (NC first so core draws on top)
            for level in LEVEL_ORDER:
                xs = [npos[k][0] for k, lab in zip(keys, labels) if lab == level and k in npos]
                ys = [npos[k][1] for k, lab in zip(keys, labels) if lab == level and k in npos]
                if xs:
                    marker = '+' if level == 'NC' else 'o'
                    kw = dict(s=size_map[level], c=LEVEL_COLORS[level],
                              marker=marker, linewidths=0.3,
                              label=legend_labels[level],
                              zorder=2 + LEVEL_ORDER.index(level))
                    if level != 'NC':
                        kw['edgecolors'] = 'gray'
                    ax.scatter(xs, ys, **kw)

            ax.set_title(title, fontsize=10)
            ax.set_aspect('equal')
            ax.axis('off')
            ax.legend(loc='lower right', fontsize=8, framealpha=0.8,
                      title='disease severity level', title_fontsize=8)

        self.fig_disease_map.tight_layout()
        self.canvas_disease_map.draw()

    # ---- EpiRank Map (Figure 7 — Tab 10) ----
    def _draw_epirank_map(self):
        """Reproduce Figure 7: EpiRank spatial distributions for 3 daytime values.

        Paper Section III.A: "Figure 7 shows the EpiRank spatial
        distributions for daytime = 0.0, 0.5, and 1.0."

        Layout: 2 rows × 3 columns.
        - Row 0 (a)(b)(c): All 353 townships (full Taiwan).
        - Row 1 (d)(e)(f): Taipei Metropolitan Area zoom (GTAIPEI_DB_IDS).
        """
        self.fig_epirank_map.clear()

        g = self.results['g']
        npos = self.results['npos']
        fig6_data = self.results['fig6_data']
        nodes_list = list(g.nodes())  # must match nx.to_numpy_array order

        size_map_tw = {'C-I': 24, 'C-II': 12, 'C-III': 6, 'NC': 2}
        size_map_tp = {'C-I': 60, 'C-II': 30, 'C-III': 15, 'NC': 6}
        legend_labels = {'NC': 'non-core', 'C-III': 'core-III',
                         'C-II': 'core-II', 'C-I': 'core-I'}
        daytimes = [0.0, 0.5, 1.0]
        top_labels = ['(a)', '(b)', '(c)']
        bot_labels = ['(d)', '(e)', '(f)']

        gs = self.fig_epirank_map.add_gridspec(2, 3, hspace=0.25, wspace=0.15,
                                                top=0.90)

        # Identify Taipei Metro nodes for zoom
        taipei_nodes = [s for s in nodes_list if g.nodes[s][KEY_DB_ID] in GTAIPEI_DB_IDS]

        for col, dt in enumerate(daytimes):
            # Build ER_rank dict from epidemic_risk matrix
            er_matrix = fig6_data[dt]['epidemic_risk']
            er_vals = np.array(er_matrix).flatten()
            er_rank = {nodes_list[i]: er_vals[i] for i in range(len(nodes_list))}

            vals_arr = np.array(list(er_rank.values()))
            keys_list = list(er_rank.keys())
            breaks = head_tail_breaks(vals_arr)
            if len(breaks) < 3:
                continue
            labels = classify_by_breaks(vals_arr, breaks)
            label_map = {k: lab for k, lab in zip(keys_list, labels)}

            # Row 0: Full Taiwan
            ax0 = self.fig_epirank_map.add_subplot(gs[0, col])
            for level in LEVEL_ORDER:
                xs = [npos[k][0] for k in keys_list if label_map[k] == level and k in npos]
                ys = [npos[k][1] for k in keys_list if label_map[k] == level and k in npos]
                if xs:
                    ax0.scatter(xs, ys, s=size_map_tw[level], c=LEVEL_COLORS[level],
                                edgecolors='gray', linewidths=0.3,
                                label=legend_labels[level],
                                zorder=2 + LEVEL_ORDER.index(level))
            ax0.set_title(f'{top_labels[col]} daytime={dt:.1f}', fontsize=10)
            ax0.set_aspect('equal')
            ax0.axis('off')
            if col == 0:
                ax0.set_ylabel('Taiwan', fontsize=10, labelpad=20)
                ax0.yaxis.set_visible(True)
                ax0.tick_params(left=False, labelleft=False)

            # Row 1: Taipei Metro zoom
            ax1 = self.fig_epirank_map.add_subplot(gs[1, col])
            for level in LEVEL_ORDER:
                xs = [npos[k][0] for k in taipei_nodes if label_map.get(k) == level and k in npos]
                ys = [npos[k][1] for k in taipei_nodes if label_map.get(k) == level and k in npos]
                if xs:
                    ax1.scatter(xs, ys, s=size_map_tp[level], c=LEVEL_COLORS[level],
                                edgecolors='gray', linewidths=0.3,
                                label=legend_labels[level],
                                zorder=2 + LEVEL_ORDER.index(level))
            # Auto zoom to Taipei region with padding
            if taipei_nodes:
                tx = [npos[k][0] for k in taipei_nodes if k in npos]
                ty = [npos[k][1] for k in taipei_nodes if k in npos]
                if tx and ty:  # 防禦：過濾後座標可能為空
                    pad_x = (max(tx) - min(tx)) * 0.08 or 0.5
                    pad_y = (max(ty) - min(ty)) * 0.08 or 0.5
                    ax1.set_xlim(min(tx) - pad_x, max(tx) + pad_x)
                    ax1.set_ylim(min(ty) - pad_y, max(ty) + pad_y)
            ax1.set_title(f'{bot_labels[col]} daytime={dt:.1f}', fontsize=10)
            ax1.set_aspect('equal')
            ax1.axis('off')
            if col == 0:
                ax1.set_ylabel('Taipei Metropolitan Area', fontsize=10, labelpad=20)
                ax1.yaxis.set_visible(True)
                ax1.tick_params(left=False, labelleft=False)

        # Shared legend at top center
        from matplotlib.patches import Patch
        legend_patches = [Patch(facecolor=LEVEL_COLORS[lv],
                                label=legend_labels[lv])
                          for lv in LEVEL_ORDER]
        self.fig_epirank_map.legend(
            handles=legend_patches, loc='upper center',
            ncol=4, fontsize=9, frameon=False,
            title='EpiRank level', title_fontsize=10,
            bbox_to_anchor=(0.5, 0.99))

        self.fig_epirank_map.subplots_adjust(left=0.05, right=0.98,
                                              bottom=0.02, top=0.90)
        self.canvas_epirank_map.draw()

    # ---- EpiRank vs Disease Map (Figure 8 — Tab 11) ----
    def _draw_overlay_map(self):
        """Reproduce Figure 8: Overlay of EpiRank core townships on actual
        disease severity in the Taipei Metropolitan Area.

        Paper Section III.C: "Figure 8 superimposes the core townships
        identified by EpiRank (black hollow circles) onto the actual
        disease severity map of the Taipei metropolitan area."

        IMPORTANT: Disease head/tail breaks must be computed GLOBALLY
        (all 353 townships), not locally on Taipei-only data, so that
        the classification matches Figure 4's global view.
        """
        self.fig_overlay_map.clear()

        g = self.results['g']
        npos = self.results['npos']
        fig6_data = self.results['fig6_data']
        nodes_list = list(g.nodes())  # must match nx.to_numpy_array order

        # ── EpiRank classification (GLOBAL, all nodes) at daytime=0.5 ──
        # 防禦：若 0.5 不在 fig6_data 中（浮點數鍵不匹配），
        # 退回使用當前 daytime 的結果。
        dt_key = 0.5 if 0.5 in fig6_data else self.results.get('daytime', list(fig6_data.keys())[0])
        er_matrix = fig6_data[dt_key]['epidemic_risk']
        er_vals = np.array(er_matrix).flatten()
        er_rank = {nodes_list[i]: er_vals[i] for i in range(len(nodes_list))}
        er_arr = np.array(list(er_rank.values()))
        er_keys = list(er_rank.keys())
        er_breaks = head_tail_breaks(er_arr)
        er_labels = classify_by_breaks(er_arr, er_breaks) if len(er_breaks) >= 3 else ['NC'] * len(er_arr)
        er_label_map = {k: lab for k, lab in zip(er_keys, er_labels)}
        er_core_nodes = {k for k, lab in er_label_map.items() if lab != 'NC'}

        # Taipei Metro nodes
        taipei_nodes = [s for s in nodes_list if g.nodes[s][KEY_DB_ID] in GTAIPEI_DB_IDS]
        if not taipei_nodes:
            self.canvas_overlay_map.draw()
            return

        # Taipei coordinate bounds
        tx = [npos[k][0] for k in taipei_nodes if k in npos]
        ty = [npos[k][1] for k in taipei_nodes if k in npos]
        if not tx or not ty:  # 防禦：無有效座標則提早返回
            self.canvas_overlay_map.draw()
            return
        pad_x = (max(tx) - min(tx)) * 0.08 or 0.5
        pad_y = (max(ty) - min(ty)) * 0.08 or 0.5

        size_map = {'C-I': 120, 'C-II': 60, 'C-III': 30, 'NC': 15}
        legend_labels = {'NC': 'non-core', 'C-III': 'core-III',
                         'C-II': 'core-II', 'C-I': 'core-I'}

        disease_info = [
            ('(a) flu cases', self.results['flu_case_rank']),
            ('(b) EV cases', self.results['ev_case_rank']),
        ]

        from matplotlib.lines import Line2D

        for col, (title, case_dict) in enumerate(disease_info):
            ax = self.fig_overlay_map.add_subplot(1, 2, col + 1)

            # ── GLOBAL classification: compute breaks on ALL 353 townships ──
            all_vals = np.array(list(case_dict.values()))
            all_keys = list(case_dict.keys())
            breaks = head_tail_breaks(all_vals)
            if len(breaks) < 3:
                ax.set_title(title + '\n(insufficient data)')
                continue
            all_labels = classify_by_breaks(all_vals, breaks)
            global_label_map = {k: lab for k, lab in zip(all_keys, all_labels)}

            # Filter to Taipei Metro nodes only for display
            taipei_keys = [k for k in taipei_nodes if k in case_dict and k in npos]

            # Background: light gray dots for all Taipei nodes (geographic context)
            bg_x = [npos[k][0] for k in taipei_keys]
            bg_y = [npos[k][1] for k in taipei_keys]
            ax.scatter(bg_x, bg_y, s=60, c='#e0e0e0', edgecolors='#d0d0d0',
                       linewidths=0.3, zorder=1)

            # Layer 1: Actual disease severity (filled circles, GLOBAL labels)
            for level in LEVEL_ORDER:
                xs = [npos[k][0] for k in taipei_keys if global_label_map.get(k) == level]
                ys = [npos[k][1] for k in taipei_keys if global_label_map.get(k) == level]
                if xs:
                    ax.scatter(xs, ys, s=size_map[level] * 1.5, c=LEVEL_COLORS[level],
                               edgecolors='gray', linewidths=0.3,
                               label=legend_labels[level],
                               zorder=2 + LEVEL_ORDER.index(level))

            # Layer 2: EpiRank core townships (black hollow circles)
            core_in_taipei = [k for k in taipei_keys if k in er_core_nodes]
            if core_in_taipei:
                cx = [npos[k][0] for k in core_in_taipei]
                cy = [npos[k][1] for k in core_in_taipei]
                ax.scatter(cx, cy, s=180, facecolors='none', edgecolors='black',
                           linewidths=2, zorder=10, label='core in daytime=0.5')

            ax.set_xlim(min(tx) - pad_x, max(tx) + pad_x)
            ax.set_ylim(min(ty) - pad_y, max(ty) + pad_y)
            ax.set_title(title, fontsize=10)
            ax.set_aspect('equal')
            ax.axis('off')
            if col == 0:
                ax.set_ylabel('Taipei Metropolitan Area', fontsize=10, labelpad=20)
                ax.yaxis.set_visible(True)
                ax.tick_params(left=False, labelleft=False)

        # Shared bottom-center legend (matching paper Figure 8)
        legend_handles = [
            Line2D([0], [0], marker='o', color='w', markerfacecolor=LEVEL_COLORS[lv],
                   markersize=8, markeredgecolor='gray', label=legend_labels[lv])
            for lv in LEVEL_ORDER
        ]
        legend_handles.append(
            Line2D([0], [0], marker='o', color='w', markerfacecolor='none',
                   markersize=10, markeredgecolor='black', markeredgewidth=2,
                   label='core in daytime=0.5'))
        self.fig_overlay_map.legend(
            handles=legend_handles, loc='lower center',
            ncol=5, fontsize=8, frameon=False,
            title='disease severity levels', title_fontsize=9,
            bbox_to_anchor=(0.5, 0.02))

        self.fig_overlay_map.tight_layout(rect=[0, 0.08, 1, 1])
        self.canvas_overlay_map.draw()

    # ---- Populate Correlations Table (Table 2 — Tab 3) ----
    def _populate_correlations(self):
        """Reproduce Table 2: Pearson/Spearman correlations, Recall, and
        Precision for EpiRank vs PageRank vs HITS vs disease cases.

        Paper Section III.B–C: "Table 2 shows the Pearson and Spearman
        correlations, recall, and precision of each index against flu
        and EV cases."

        Layout matches the paper:
        Rows = Disease × Statistic, Columns = EpiRank | PageRank | HITS-Hub | HITS-Authority
        Paper layout:
          Disease | Index        | EpiRank | PageRank | HITS-Hub | HITS-Authority
          Flu     | Pearson's R  |  0.513  |  0.355   |  0.387   |  0.397
                  | Spearman Rho |  0.531  |  0.445   |  0.400   |  0.403
                  | Recall       |  0.724  |  0.609   |  0.299   |  0.333
                  | Precision    |  0.553  |  0.495   |  1.000   |  1.000
          EV      | ...
        """
        corr = self.results['correlations']

        # Disease case dicts
        disease_info = [
            ('Flu', self.results['flu_case_rank']),
            ('EV',  self.results['ev_case_rank']),
        ]
        # Network index dicts  (display_name, rank_dict, corr_key)
        index_info = [
            ('EpiRank',        self.results['ER_rank'],        'EpiRank'),
            ('PageRank',       self.results['page_rank'],      'PageRank'),
            ('HITS-Hub',       self.results['hub_rank'],       'HITS Hub'),
            ('HITS-Authority', self.results['authority_rank'], 'HITS Authority'),
        ]
        stat_names = ["Pearson's R", "Spearman's Rho", "Recall", "Precision"]

        # Pre-compute all values: data[disease][index] = (pearson, spearman, recall, precision)
        data = {}
        for disease_name, disease_dict in disease_info:
            disease_vals = np.array(list(disease_dict.values()))
            disease_keys = list(disease_dict.keys())
            disease_breaks = head_tail_breaks(disease_vals)
            if len(disease_breaks) < 3:
                continue
            disease_labels = classify_by_breaks(disease_vals, disease_breaks)
            actual_core = {k for k, lab in zip(disease_keys, disease_labels) if lab != 'NC'}

            data[disease_name] = {}
            for display_name, index_dict, corr_key in index_info:
                common_keys = sorted(set(index_dict.keys()) & set(disease_dict.keys()))
                idx_vals = np.array([index_dict[k] for k in common_keys])
                idx_breaks = head_tail_breaks(idx_vals)
                if len(idx_breaks) < 3:
                    data[disease_name][display_name] = (float('nan'),) * 4
                    continue
                idx_labels = classify_by_breaks(idx_vals, idx_breaks)
                predicted_core = {k for k, lab in zip(common_keys, idx_labels) if lab != 'NC'}

                tp = len(actual_core & predicted_core)
                fp = len(predicted_core - actual_core)
                fn = len(actual_core - predicted_core)
                recall    = tp / (tp + fn) if (tp + fn) > 0 else 0.0
                precision = tp / (tp + fp) if (tp + fp) > 0 else 0.0

                pearson_r = spearman_rho = float('nan')
                if corr_key in corr and disease_name in corr[corr_key]:
                    pearson_r    = corr[corr_key][disease_name]['Pearson'][0]
                    spearman_rho = corr[corr_key][disease_name]['Spearman'][0]

                data[disease_name][display_name] = (pearson_r, spearman_rho, recall, precision)

        # Build the table: rows = Disease × Statistic, cols = indexes
        index_names = [info[0] for info in index_info]
        headers = ['Disease', 'Index'] + index_names
        n_stats = len(stat_names)
        diseases_with_data = [d for d in ['Flu', 'EV'] if d in data]
        total_rows = len(diseases_with_data) * n_stats

        self.corr_table.setRowCount(total_rows)
        self.corr_table.setColumnCount(len(headers))
        self.corr_table.setHorizontalHeaderLabels(headers)

        row = 0
        for disease_name in diseases_with_data:
            for s_idx, stat_name in enumerate(stat_names):
                # Disease column — show name only on first stat row
                disease_item = QTableWidgetItem(disease_name if s_idx == 0 else '')
                if s_idx == 0:
                    disease_item.setFont(QFont('Arial', weight=QFont.Bold))
                self.corr_table.setItem(row, 0, disease_item)

                # Statistic name column
                stat_item = QTableWidgetItem(stat_name)
                self.corr_table.setItem(row, 1, stat_item)

                # Value columns for each index
                for c_idx, idx_name in enumerate(index_names):
                    vals = data[disease_name].get(idx_name, (float('nan'),) * 4)
                    val = vals[s_idx]
                    if np.isnan(val):
                        cell = QTableWidgetItem('N/A')
                    else:
                        cell = QTableWidgetItem(f"{val:.3f}")
                    cell.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                    self.corr_table.setItem(row, 2 + c_idx, cell)

                row += 1

        self.corr_table.resizeColumnsToContents()
        for col in range(len(headers)):
            w = self.corr_table.columnWidth(col)
            self.corr_table.setColumnWidth(col, max(w, 110))

    # ---- Auto-save results (matching original ERA.py output format) ----
    def _auto_save_results(self):
        """
        Automatically save 3 output files to the data directory,
        matching the old ERA.py behavior:
          1. ERA_result_d_{d}_daytime_{daytime}_loops_{iters}.xlsx  (Excel results)
          2. ERA_result_d_{d}_daytime_{daytime}_loops_{iters}.png   (Network map)
          3. ERA_result.net                                          (Pajek network)
        """
        r = self.results
        d = r['d']
        daytime = r['daytime']
        iters = r['iterations']
        base_name = f"ERA_result_d_{d}_daytime_{daytime}_loops_{iters}"

        try:
            # --- 1) Save .xlsx ---
            xlsx_path = os.path.join(self.data_dir, base_name + '.xlsx')
            self._save_excel_to_path(xlsx_path)
            self._log(f"Auto-saved: {xlsx_path}")

            # --- 2) Save .png (network map) ---
            png_path = os.path.join(self.data_dir, base_name + '.png')
            self.fig_network.savefig(png_path, dpi=450, bbox_inches='tight',
                                      facecolor='white')
            self._log(f"Auto-saved: {png_path}")

            # --- 3) Save .net (Pajek format) ---
            # NetworkX 3.x write_pajek drops non-string attributes.
            # Convert all node/edge attributes to strings for compatibility.
            g_pajek = r['g'].copy()
            for n in g_pajek.nodes():
                for attr_key in list(g_pajek.nodes[n].keys()):
                    g_pajek.nodes[n][attr_key] = str(g_pajek.nodes[n][attr_key])
            for u, v, data in g_pajek.edges(data=True):
                for attr_key in list(data.keys()):
                    data[attr_key] = str(data[attr_key])
            net_path = os.path.join(self.data_dir, 'ERA_result.net')
            nx.write_pajek(g_pajek, net_path)
            del g_pajek
            self._log(f"Auto-saved: {net_path}")

            self._log(f"\n--- All 3 output files saved to: {self.data_dir} ---")
        except Exception as e:
            self._log(f"\nWARNING: Auto-save error: {e}")

    def _save_excel_to_path(self, path):
        """Save results to Excel at the given path.

        Used by both ``_auto_save_results()`` and manual ``_export_excel()``.
        The spreadsheet layout is inherited from the original ERA.py and
        uses hard-coded column positions:

        Excel column layout (1-indexed):
        ┌─────────────────────────────────────────────────────────────┐
        │  Col A (1)     : Network info text (rows 1-14)             │
        │  Col B–AC (2–29): Main data table (row 1 = headers)        │
        │    B(2)  seq no        Q(17) C.local                       │
        │    C(3)  post code     R(18) C.out                         │
        │    D(4)  db ID         S(19) C.in                          │
        │    E(5)  county        T(20) C.out-towns (in-degree)       │
        │    F(6)  town          U(21) C.in-towns  (out-degree)      │
        │    G(7)  pos.x         V(22) R.zone                       │
        │    H(8)  pos.y         W(23) page rank                     │
        │    I(9)  population    X(24) hits.hub                      │
        │    J(10) area          Y(25) hits.authority                │
        │    K(11) density       Z(26) flu cases                     │
        │    L(12) normalized D  AA(27) ev cases                     │
        │    M(13) age 0~14      AB(28) sars cases                   │
        │    N(14) age 15~64                                         │
        │    O(15) age 65+                                           │
        │    P(16) ERV                                               │
        ├─────────────────────────────────────────────────────────────┤
        │  Risk classification sidebar (row 1 = threshold labels):   │
        │    AD-AE (30-31): AVG+2STD (very high)  — county, town     │
        │    AG-AH (33-34): AVG+STD  (high)       — county, town     │
        │    AJ-AK (36-37): AVG      (medium)     — county, town     │
        │    AM-AN (39-40): AVG-STD  (low)        — county, town     │
        │    AP-AQ (42-43): AVG-2STD (very low)   — county, town     │
        ├─────────────────────────────────────────────────────────────┤
        │  Statistics section (starts at row = 353 + 2 = 355):       │
        │    Col H-I  (8-9) : Population vs ER/Flu/SARS correlations │
        │    Col O-P (15-16): ER total / mean / SD                   │
        │    Col W    (23)  : Correlation row labels                 │
        │    Col X    (24)  : PageRank vs EV/Flu/SARS correlations   │
        │    Col Y    (25)  : HITS-Hub vs EV/Flu/SARS correlations   │
        │    Col Z    (26)  : HITS-Auth vs EV/Flu/SARS correlations  │
        │    Col AA   (27)  : EpiRank vs Flu correlations            │
        │    Col AB   (28)  : EpiRank vs EV correlations             │
        │    Col AC   (29)  : EpiRank vs SARS correlations (gTaipei) │
        │  Note: SARS correlations use gTaipei subset only.          │
        └─────────────────────────────────────────────────────────────┘
        """
        r = self.results
        wb = Workbook()
        ws = wb.active
        ws.title = "EpiRank Results"

        g = r['g']
        town_data = r['town_data']
        ER_avg = r['ER_avg']
        ER_std = r['ER_std']
        ER_tot = r['ER_tot']

        # Info section (matching old ERA.py layout)
        info_font = Font(name='Arial Narrow', color='8B0000', bold=True)
        title_font = Font(name='Arial Narrow', color='8B0000', bold=True)
        body_font = Font(name='Arial Narrow', color='00008B')

        in_degrees = dict(g.in_degree())
        out_degrees = dict(g.out_degree())
        in_vals = list(in_degrees.values())
        out_vals = list(out_degrees.values())

        info_rows = [
            f"number of nodes = {g.number_of_nodes()}",
            f"number of edges = {g.number_of_edges()}",
            f"average in-degree = {np.mean(in_vals)}",
            f"STD of in-degree = {np.std(in_vals)}",
            f"average out-degree = {np.mean(out_vals)}",
            f"STD of out-degree = {np.std(out_vals)}",
            "", "", "", "",
            f"number of selfloop edges = {nx.number_of_selfloops(g)}",
            f"number of epidemic analysis loops = {r['iterations']}",
            f"parameter.d = {round(r['d'], 2)}",
            f"parameter.daytime = {round(r['daytime'], 2)}",
        ]
        for i, text in enumerate(info_rows):
            ws.cell(row=i + 1, column=1, value=text).font = info_font

        # Column headers (row 1, starting col B = col 2)
        headers = ['seq no', 'post code', 'db ID', 'county', 'town',
                    'pos.x', 'pos.y', 'population', 'area (km^2)',
                    'density (D)', 'normalized D',
                    'age  0 ~ 14', 'age 15 ~ 64', 'age 65+',
                    'ERV', 'ERP (%)',
                    'C.local', 'C.out', 'C.in', 'C.out-towns', 'C.in-towns',
                    'R.zone', 'page rank', 'hits.hub', 'hits.authority',
                    'flu cases', 'ev cases', 'sars cases']
        for j, h in enumerate(headers):
            cell = ws.cell(row=1, column=j + 2, value=h)
            cell.font = title_font
            cell.alignment = Alignment(horizontal='center')

        # Threshold labels
        ws.cell(row=1, column=30, value='AVG+2STD =').font = info_font
        ws.cell(row=1, column=31, value=round(ER_avg + 2 * ER_std, 5)).font = info_font
        ws.cell(row=1, column=33, value='AVG+STD =').font = info_font
        ws.cell(row=1, column=34, value=round(ER_avg + ER_std, 5)).font = info_font
        ws.cell(row=1, column=36, value='AVG=').font = info_font
        ws.cell(row=1, column=37, value=round(ER_avg, 5)).font = info_font
        ws.cell(row=1, column=39, value='AVG-STD =').font = info_font
        ws.cell(row=1, column=40, value=round(ER_avg - ER_std, 5)).font = info_font
        ws.cell(row=1, column=42, value='AVG-2STD =').font = info_font
        ws.cell(row=1, column=43, value=round(ER_avg - 2 * ER_std, 5)).font = info_font

        # Data rows (all 353 towns, sorted by ERV descending)
        row_idx = 2  # data starts at row 2
        # Risk classification sidebar columns (matching old ERA.py layout)
        # Each risk level writes county/town pairs into its own Excel column.
        risk_keys = ['vh', 'h', 'm', 'l', 'vl']
        risk_cols = {'vh': 30, 'h': 33, 'l': 39, 'm': 36, 'vl': 42}
        risk_row_idx = {k: 2 for k in risk_keys}

        for item in r['table_data']:
            seq_no = item['seq_no']
            db_ID = item['db_ID']
            td = town_data[db_ID]
            erv = item['ERV']

            vals = [
                seq_no, item['post_code'], db_ID,
                td[KEY_COUNTY], td[KEY_TOWN],
                td[KEY_POS_XY][0], td[KEY_POS_XY][1],
                td[KEY_POPULATION],
                round(td[KEY_AREA], 2), round(td[KEY_DENSITY], 2),
                round(td[KEY_NORMALIZED_DENSITY], 6),
                str(round(td[KEY_AGE_0_14], 2)) + '%',
                str(round(td[KEY_AGE_15_64], 2)) + '%',
                str(round(td[KEY_AGE_65], 2)) + '%',
                round(erv, 5),
                str(round(100.0 * erv / ER_tot, 2)) + '%' if ER_tot > 0 else '0%',
                td[KEY_LOCAL_COMMUTER_TYPE1],
                td[KEY_OUT_COMMUTER_TYPE1], td[KEY_IN_COMMUTER_TYPE1],
                g.in_degree(seq_no), g.out_degree(seq_no),
                td.get('railroad_zone', 0),
                round(r['page_rank'].get(seq_no, 0), 6),
                round(r['hub_rank'].get(seq_no, 0), 6),
                round(r['authority_rank'].get(seq_no, 0), 6),
                td.get(KEY_FLU_TOTAL_CASES, 0),
                td.get(KEY_EV_AVERAGE_CASES, 0),
                td.get(KEY_SARS_TOTAL_CASES, 0),
            ]
            for j, v in enumerate(vals):
                ws.cell(row=row_idx, column=j + 2, value=v).font = body_font

            # Risk classification sidebar (matching old ERA.py columns 29-43)
            if erv >= ER_avg + 2 * ER_std:
                key = 'vh'
            elif erv >= ER_avg + ER_std:
                key = 'h'
            elif erv >= ER_avg:
                key = 'm'
            elif erv >= ER_avg - ER_std:
                key = 'l'
            else:
                key = 'vl'
            c = risk_cols[key]
            ri = risk_row_idx[key]
            ws.cell(row=ri, column=c, value=td[KEY_COUNTY]).font = body_font
            ws.cell(row=ri, column=c + 1, value=td[KEY_TOWN]).font = body_font
            risk_row_idx[key] += 1

            row_idx += 1

        # Statistics and correlations section (after data, matching old ERA.py)
        sr = row_idx  # stats start row
        ER_rank = r['ER_rank']
        pop_rank = r['pop_rank']
        flu_case_rank = r['flu_case_rank']
        ev_case_rank = r['ev_case_rank']
        sars_case_rank = r['sars_case_rank']
        page_rank = r['page_rank']
        hub_rank = r['hub_rank']
        authority_rank = r['authority_rank']

        # gTaipei subsets (same as old ERA.py)
        nodes_list = list(g.nodes())
        gTaipei_ER_rank = {s: ER_rank[s] for s in nodes_list if g.nodes[s][KEY_DB_ID] in GTAIPEI_DB_IDS}
        gTaipei_sars = {s: sars_case_rank[s] for s in nodes_list if g.nodes[s][KEY_DB_ID] in GTAIPEI_DB_IDS}
        gTaipei_pr_ids = set(range(0, 29)) | set(range(303, 310)) | set(range(330, 342))
        gTaipei_page = {k: v for k, v in page_rank.items() if k in gTaipei_pr_ids}
        gTaipei_hub = {k: v for k, v in hub_rank.items() if k in gTaipei_pr_ids}
        gTaipei_auth = {k: v for k, v in authority_rank.items() if k in gTaipei_pr_ids}

        # Population vs EpiRank correlations
        ws.cell(row=sr, column=8, value='pearson = ').font = body_font
        ws.cell(row=sr + 2, column=8, value='spearman = ').font = body_font
        ws.cell(row=sr, column=9, value=get_pearson_cor(pop_rank, ER_rank)[0]).font = body_font
        ws.cell(row=sr + 2, column=9, value=get_spearman_cor(pop_rank, ER_rank)[0]).font = body_font
        ws.cell(row=sr + 4, column=9, value=get_pearson_cor(pop_rank, flu_case_rank)[0]).font = body_font
        ws.cell(row=sr + 6, column=9, value=get_spearman_cor(pop_rank, flu_case_rank)[0]).font = body_font
        ws.cell(row=sr + 8, column=9, value=get_pearson_cor(pop_rank, sars_case_rank)[0]).font = body_font
        ws.cell(row=sr + 10, column=9, value=get_spearman_cor(pop_rank, sars_case_rank)[0]).font = body_font

        # ER stats
        ws.cell(row=sr, column=15, value='total = ').font = body_font
        ws.cell(row=sr, column=16, value=round(ER_tot, 5)).font = body_font
        ws.cell(row=sr + 1, column=15, value='mean = ').font = body_font
        ws.cell(row=sr + 1, column=16, value=round(ER_avg, 5)).font = body_font
        ws.cell(row=sr + 2, column=15, value='SD   = ').font = body_font
        ws.cell(row=sr + 2, column=16, value=round(ER_std, 5)).font = body_font

        # Correlation labels
        labels = [
            'pearson  (*,  EV)', 'spearman (*,  EV)',
            'pearson  (*, Flu)', 'spearman (*, Flu)',
            'pearson  (*, SARS)', 'spearman (*, SARS)', 'kendall. (*, SARS)',
        ]
        for i, lbl in enumerate(labels):
            ws.cell(row=sr + i, column=23, value=lbl).font = body_font

        # PageRank correlations (col 24)
        ws.cell(row=sr, column=24, value=get_pearson_cor(page_rank, ev_case_rank)[0]).font = body_font
        ws.cell(row=sr + 1, column=24, value=get_spearman_cor(page_rank, ev_case_rank)[0]).font = body_font
        ws.cell(row=sr + 2, column=24, value=get_pearson_cor(page_rank, flu_case_rank)[0]).font = body_font
        ws.cell(row=sr + 3, column=24, value=get_spearman_cor(page_rank, flu_case_rank)[0]).font = body_font
        ws.cell(row=sr + 4, column=24, value=get_pearson_cor(gTaipei_page, gTaipei_sars)[0]).font = body_font
        ws.cell(row=sr + 5, column=24, value=get_spearman_cor(gTaipei_page, gTaipei_sars)[0]).font = body_font
        ws.cell(row=sr + 6, column=24, value=get_kendalltau_cor(gTaipei_page, gTaipei_sars)[0]).font = body_font

        # HITS Hub correlations (col 25)
        ws.cell(row=sr, column=25, value=get_pearson_cor(hub_rank, ev_case_rank)[0]).font = body_font
        ws.cell(row=sr + 1, column=25, value=get_spearman_cor(hub_rank, ev_case_rank)[0]).font = body_font
        ws.cell(row=sr + 2, column=25, value=get_pearson_cor(hub_rank, flu_case_rank)[0]).font = body_font
        ws.cell(row=sr + 3, column=25, value=get_spearman_cor(hub_rank, flu_case_rank)[0]).font = body_font
        ws.cell(row=sr + 4, column=25, value=get_pearson_cor(gTaipei_hub, gTaipei_sars)[0]).font = body_font
        ws.cell(row=sr + 5, column=25, value=get_spearman_cor(gTaipei_hub, gTaipei_sars)[0]).font = body_font
        ws.cell(row=sr + 6, column=25, value=get_kendalltau_cor(gTaipei_hub, gTaipei_sars)[0]).font = body_font

        # HITS Authority correlations (col 26)
        ws.cell(row=sr, column=26, value=get_pearson_cor(authority_rank, ev_case_rank)[0]).font = body_font
        ws.cell(row=sr + 1, column=26, value=get_spearman_cor(authority_rank, ev_case_rank)[0]).font = body_font
        ws.cell(row=sr + 2, column=26, value=get_pearson_cor(authority_rank, flu_case_rank)[0]).font = body_font
        ws.cell(row=sr + 3, column=26, value=get_spearman_cor(authority_rank, flu_case_rank)[0]).font = body_font
        ws.cell(row=sr + 4, column=26, value=get_pearson_cor(gTaipei_auth, gTaipei_sars)[0]).font = body_font
        ws.cell(row=sr + 5, column=26, value=get_spearman_cor(gTaipei_auth, gTaipei_sars)[0]).font = body_font
        ws.cell(row=sr + 6, column=26, value=get_kendalltau_cor(gTaipei_auth, gTaipei_sars)[0]).font = body_font

        # EpiRank vs Flu (col 27), EV (col 28), SARS (col 29)
        ws.cell(row=sr, column=28, value=get_pearson_cor(ER_rank, ev_case_rank)[0]).font = body_font
        ws.cell(row=sr + 1, column=28, value=get_spearman_cor(ER_rank, ev_case_rank)[0]).font = body_font
        ws.cell(row=sr + 2, column=27, value=get_pearson_cor(ER_rank, flu_case_rank)[0]).font = body_font
        ws.cell(row=sr + 3, column=27, value=get_spearman_cor(ER_rank, flu_case_rank)[0]).font = body_font
        ws.cell(row=sr + 4, column=29, value=get_pearson_cor(gTaipei_ER_rank, gTaipei_sars)[0]).font = body_font
        ws.cell(row=sr + 5, column=29, value=get_spearman_cor(gTaipei_ER_rank, gTaipei_sars)[0]).font = body_font
        ws.cell(row=sr + 6, column=29, value=get_kendalltau_cor(gTaipei_ER_rank, gTaipei_sars)[0]).font = body_font

        # Column labels row
        for c, lbl in [(24, 'page rank'), (25, 'hits.hub'), (26, 'hits.authority'),
                         (27, 'flu cases'), (28, 'ev cases'), (29, 'sars cases')]:
            ws.cell(row=sr + 7, column=c, value=lbl).font = title_font

        wb.save(path)

    # ---- Sensitivity Analysis (Figure 11 from paper) ----
    def _run_sensitivity(self):
        """Launch sensitivity analysis in a background thread."""
        if not self.results:
            QMessageBox.information(self, "No Data", "Please run EpiRank first.")
            return

        self.btn_run.setEnabled(False)
        self.btn_sensitivity.setEnabled(False)
        self.progress_bar.setValue(0)
        self._log("\n=== Starting Sensitivity Analysis ===")
        self._log("Computing EpiRank for 21 daytime × 20 damping factor = 420 combinations...")

        self.sensitivity_worker = SensitivityWorker(
            self.data_dir,
            self.spin_loops.value()
        )
        self.sensitivity_worker.progress.connect(self._on_sensitivity_progress)
        self.sensitivity_worker.log_message.connect(self._log)
        self.sensitivity_worker.finished_ok.connect(self._on_sensitivity_ok)
        self.sensitivity_worker.finished_err.connect(self._on_sensitivity_err)
        self.sensitivity_worker.start()
        self._update_status("Running sensitivity analysis...")

    def _on_sensitivity_progress(self, cur, total):
        if total > 0:
            self.progress_bar.setValue(int(100 * cur / total))

    def _on_sensitivity_ok(self, results):
        self.btn_run.setEnabled(True)
        self.btn_sensitivity.setEnabled(True)
        self.progress_bar.setValue(100)
        self.sensitivity_results = results
        self._log("Sensitivity analysis completed successfully.")
        self._update_status("Sensitivity analysis completed.")
        self._draw_sensitivity()
        self.tabs.setCurrentIndex(13)  # Switch to Sensitivity Analysis tab

    def _on_sensitivity_err(self, err_msg):
        self.btn_run.setEnabled(True)
        self.btn_sensitivity.setEnabled(True)
        self.progress_bar.setValue(0)
        self._log(f"\nSensitivity Analysis ERROR:\n{err_msg}")
        self._update_status("Sensitivity analysis error.")
        QMessageBox.critical(self, "Sensitivity Analysis Error", str(err_msg)[:500])

    def _draw_sensitivity(self):
        """Reproduce Figure 11: 2×2 heatmaps of correlation vs (daytime, d).

        Paper Section III.D: "Figure 11 shows the sensitivity analysis
        of correlation between EpiRank and actual disease data for
        different values of the daytime parameter and damping factor."

        Subplots:
          (a) Flu Pearson's R, (b) Flu Spearman's ρ,
          (c) EV Pearson's R,  (d) EV Spearman's ρ.
        """
        self.fig_sensitivity.clear()

        sr = self.sensitivity_results
        daytime_vals = sr['daytime_values']
        d_vals = sr['d_values']

        matrices = [
            ("(a) flu (Pearson's R)", sr['flu_pearson']),
            ("(b) flu (Spearman's Rho)", sr['flu_spearman']),
            ("(c) EV (Pearson's R)", sr['ev_pearson']),
            ("(d) EV (Spearman's Rho)", sr['ev_spearman']),
        ]

        for idx, (title, matrix) in enumerate(matrices):
            ax = self.fig_sensitivity.add_subplot(2, 2, idx + 1)

            # origin='upper' so daytime 0.0 is at top, 1.0 at bottom (matching paper)
            im = ax.imshow(matrix, aspect='auto', origin='upper',
                           cmap='GnBu',
                           extent=[d_vals[0] - 0.025, d_vals[-1] + 0.025,
                                   daytime_vals[-1] + 0.025, daytime_vals[0] - 0.025])

            # Colorbar
            cbar = self.fig_sensitivity.colorbar(im, ax=ax, fraction=0.046, pad=0.04)
            cbar.ax.tick_params(labelsize=8)

            ax.set_xlabel('damping factor', fontsize=9)
            ax.set_ylabel('daytime parameter', fontsize=9)
            ax.set_title(title, fontsize=10)

            # Tick labels — every 0.05 for x-axis (matching paper)
            x_ticks = np.arange(0.05, 1.025, 0.05)
            y_ticks = np.arange(0.0, 1.05, 0.1)
            ax.set_xticks(x_ticks)
            ax.set_yticks(y_ticks)
            ax.set_xticklabels([f'{v:.2f}' for v in x_ticks], fontsize=5, rotation=90)
            ax.set_yticklabels([f'{v:.1f}' for v in y_ticks], fontsize=7)

        self.fig_sensitivity.tight_layout()
        self.canvas_sensitivity.draw()

    # ---- Export to Excel (manual, via File menu) ----
    def _export_excel(self):
        if not self.results:
            QMessageBox.information(self, "No Data", "Please run EpiRank first.")
            return

        path, _ = QFileDialog.getSaveFileName(self, "Export Results",
                    f"ERA_result_d_{self.results['d']}_daytime_{self.results['daytime']}.xlsx",
                    "Excel Files (*.xlsx)")
        if not path:
            return

        try:
            self._save_excel_to_path(path)
            self._log(f"\nResults exported to: {path}")
            self._update_status(f"Exported to {path}")
            QMessageBox.information(self, "Export Complete", f"Results saved to:\n{path}")
        except Exception as e:
            self._log(f"\nExport error: {e}")
            QMessageBox.critical(self, "Export Error", str(e)[:500])

    # ---- Save Figure ----
    def _save_figure(self):
        """Save the current chart tab's figure to PNG/PDF/SVG via dialog."""
        tab_idx = self.tabs.currentIndex()
        fig_map = {
            1: (self.fig_network, "network_map"),
            4: (self.fig_analysis, "commuter_flow"),
            5: (self.fig_disease, "frequency_distributions"),
            6: (self.fig_epirank_dist, "frequency_distribution"),
            7: (self.fig_epirank_vs, "epirank_vs_disease"),
            8: (self.fig_index_comp, "index_comparison"),
            9: (self.fig_disease_map, "disease_map"),
            10: (self.fig_epirank_map, "epirank_map"),
            11: (self.fig_overlay_map, "epirank_vs_disease_map"),
            13: (self.fig_sensitivity, "sensitivity_analysis"),
        }
        if tab_idx not in fig_map:
            QMessageBox.information(self, "Save Chart", "Please switch to a chart tab first.")
            return

        fig, default_name = fig_map[tab_idx]
        path, _ = QFileDialog.getSaveFileName(self, "Save Chart",
                    f"{default_name}.png",
                    "PNG (*.png);;PDF (*.pdf);;SVG (*.svg)")
        if path:
            try:
                fig.savefig(path, dpi=300, bbox_inches='tight')
                self._log(f"Chart saved to: {path}")
                self._update_status(f"Chart saved to {path}")
            except Exception as e:
                self._log(f"Error saving chart: {e}")
                QMessageBox.critical(self, "Save Error", str(e)[:500])


# ============================================================
# Entry Point
# ============================================================
def main():
    # 防禦：避免在 Jupyter / IPython 等環境中重複建立 QApplication
    app = QApplication.instance() or QApplication(sys.argv)
    app.setStyle('Fusion')

    # Set default font that supports CJK
    font = QFont()
    font.setPointSize(12)
    app.setFont(font)

    window = EpiRankMainWindow()
    window.show()
    sys.exit(app.exec())


if __name__ == '__main__':
    main()
